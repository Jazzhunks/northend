"""Iteration 10: Bulk Scholarship Registration - empty rows / duplicates / missing fields."""
import io
import os
import time
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
SID = "1ed94009-d949-4504-b291-68e3571a5a44"  # WATH campaign

HEADERS = ["full_name", "email", "phone", "class", "school_institute", "venue"]


@pytest.fixture(scope="module")
def admin_session():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@northend.edu", "password": "Admin@2025"})
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    body = r.json()
    tok = body.get("access_token") or body.get("token")
    if tok:
        s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


def _mk_xlsx(rows):
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _mk_xlsx_with_empty_trailing(rows, empty_count):
    """Manually append empty rows via row_dimensions/max_row bump."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    # Force trailing empty rows by writing empty strings across all columns
    start = ws.max_row + 1
    for i in range(empty_count):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=start + i, column=col, value="")
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _unique_email(prefix):
    return f"bulk_{prefix}_{int(time.time()*1000)}@test.com"


def _unique_phone():
    return str(int(time.time() * 1000))[-10:]


# --- Template download tests ---

def test_template_download_requires_auth():
    r = requests.get(f"{BASE_URL}/api/admin/scholarships/{SID}/bulk-register-template")
    assert r.status_code == 401, f"expected 401 without auth, got {r.status_code}"


def test_template_download_success(admin_session):
    r = admin_session.get(f"{BASE_URL}/api/admin/scholarships/{SID}/bulk-register-template")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == HEADERS, f"template headers mismatch: {headers}"
    assert ws.max_row >= 2, "template should include at least one sample row"
    sample = [c.value for c in ws[2]]
    assert sample[0] and sample[1] and sample[2], f"sample row incomplete: {sample}"


# --- Bulk register tests ---

def test_bulk_register_valid_rows_with_trailing_empty(admin_session):
    e1, e2 = _unique_email("a"), _unique_email("b")
    p1, p2 = _unique_phone(), str(int(_unique_phone()) + 1)
    xlsx = _mk_xlsx_with_empty_trailing(
        [
            ["Test A", e1, p1, "Class 10", "DPS", "90 FT"],
            ["Test B", e2, p2, "Class 10", "DPS", "90 FT"],
        ],
        empty_count=500,
    )
    r = admin_session.post(
        f"{BASE_URL}/api/admin/scholarships/{SID}/bulk-register",
        files={"file": ("bulk.xlsx", xlsx.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, f"{r.status_code} {r.text}"
    body = r.json()
    print("EMPTY-TRAILING RESPONSE:", body)
    assert body["registered"] == 2, body
    assert body["skipped"] == 0, body
    assert body["total_rows"] == 2, body
    assert body["errors"] == [], body


def test_bulk_register_duplicate_row(admin_session):
    e = _unique_email("dup")
    p = _unique_phone()
    xlsx = _mk_xlsx([
        ["Dup A", e, p, "Class 10", "DPS", "90 FT"],
        ["Other", _unique_email("other"), str(int(p) + 1), "Class 10", "DPS", "90 FT"],
        ["Dup A2", e, p, "Class 10", "DPS", "90 FT"],  # dup of row 1
    ])
    r = admin_session.post(
        f"{BASE_URL}/api/admin/scholarships/{SID}/bulk-register",
        files={"file": ("bulk.xlsx", xlsx.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    print("DUP RESPONSE:", body)
    assert body["registered"] == 2, body
    assert body["skipped"] == 1, body
    assert body["total_rows"] == 3, body
    assert len(body["errors"]) == 1
    err = body["errors"][0]
    assert err.get("reason") == "duplicate", err
    assert err.get("application_no"), f"duplicate should include existing app_no: {err}"


def test_bulk_register_missing_required(admin_session):
    xlsx = _mk_xlsx([
        ["No Email", "", _unique_phone(), "Class 10", "DPS", "90 FT"],
    ])
    r = admin_session.post(
        f"{BASE_URL}/api/admin/scholarships/{SID}/bulk-register",
        files={"file": ("bulk.xlsx", xlsx.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    print("MISSING RESPONSE:", body)
    assert body["registered"] == 0, body
    assert body["skipped"] == 1, body
    assert len(body["errors"]) == 1
    assert body["errors"][0].get("reason") == "missing_required"
