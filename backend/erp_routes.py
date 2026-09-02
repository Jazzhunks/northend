from __future__ import annotations
import io
import os
import json
import uuid
import asyncio
import openpyxl
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query, Response, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr, Field

from erp_pdf import fee_receipt_pdf

# -- Constants
ROLES_ALL = {"super_admin", "center_manager", "accountant", "counsellor"}
ROLES_BRANCH = {"center_manager", "accountant", "counsellor"}
EXPENSE_CATEGORIES = ["Salary", "Rent", "Electricity", "Internet", "Marketing", "Maintenance", "Miscellaneous"]
PAYMENT_MODES = ["cash", "upi", "online", "cheque", "card"]
LEAD_STATUSES = ["new", "contacted", "follow_up", "converted", "lost"]

CGST_RATE = 9.0
SGST_RATE = 9.0

# Broadcasters Memory Matrix Mapping Layer
branch_broadcast_queues: Dict[str, List[asyncio.Queue]] = {}

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def new_id():
    return str(uuid.uuid4())

# ====== PYDANTIC COMPLIANCE VALIDATORS ======
class StaffCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: Literal["center_manager", "accountant", "counsellor"]
    branch_id: str
    phone: Optional[str] = None

class StaffUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[Literal["center_manager", "accountant", "counsellor"]] = None
    branch_id: Optional[str] = None
    phone: Optional[str] = None
    active: Optional[bool] = None
    new_password: Optional[str] = None

class BranchUpdate(BaseModel):
    gstin: Optional[str] = None
    signatory_name: Optional[str] = None
    state_code: Optional[str] = None
    manager_user_id: Optional[str] = None

class StudentCreate(BaseModel):
    full_name: str
    gender: Optional[str] = None
    dob: Optional[str] = None
    school_institute: Optional[str] = None
    board: Optional[str] = None
    category: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_email: Optional[EmailStr] = None
    emergency_phone: Optional[str] = None
    contact_phone: str
    contact_email: Optional[EmailStr] = None
    address: Optional[str] = None
    photo_url: Optional[str] = None
    course_id: str
    batch: Optional[str] = None
    branch_id: str
    counsellor_id: Optional[str] = None
    admission_date: Optional[str] = None
    scholarship_percent: float = 0.0
    discount: float = 0.0
    total_fee: float
    documents: List[dict] = Field(default_factory=list)
    notes: Optional[str] = None
    public_user_id: Optional[str] = None

class StudentUpdate(BaseModel):
    full_name: Optional[str] = None
    gender: Optional[str] = None
    dob: Optional[str] = None
    school_institute: Optional[str] = None
    board: Optional[str] = None
    category: Optional[str] = None
    parent_name: Optional[str] = None
    parent_phone: Optional[str] = None
    parent_email: Optional[EmailStr] = None
    emergency_phone: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    address: Optional[str] = None
    photo_url: Optional[str] = None
    batch: Optional[str] = None
    counsellor_id: Optional[str] = None
    scholarship_percent: Optional[float] = None
    discount: Optional[float] = None
    total_fee: Optional[float] = None
    documents: Optional[List[dict]] = None
    notes: Optional[str] = None
    status: Optional[Literal["active", "inactive", "alumni"]] = None

class PaymentCreate(BaseModel):
    student_id: str
    amount: float
    mode: Literal["cash", "upi", "online", "cheque", "card"]
    next_due_date: Optional[str] = None
    notes: Optional[str] = None
    transaction_ref: Optional[str] = None
    apply_gst: bool = True

class ExpenseCreate(BaseModel):
    branch_id: str
    category: Literal["Salary", "Rent", "Electricity", "Internet", "Marketing", "Maintenance", "Miscellaneous"]
    amount: float
    description: str
    vendor: Optional[str] = None
    bill_url: Optional[str] = None
    expense_date: Optional[str] = None

class ExpenseDecision(BaseModel):
    decision: Literal["approve", "reject"]
    note: Optional[str] = None

class LeadCreate(BaseModel):
    name: str
    phone: str
    email: Optional[EmailStr] = None
    gender: Optional[str] = None
    dob: Optional[str] = None
    school_institute: Optional[str] = None
    city: Optional[str] = None
    target_exam: Optional[str] = None
    preferred_batch: Optional[str] = None
    source: Optional[str] = None
    branch_id: str
    counsellor_id: Optional[str] = None
    notes: Optional[str] = None

class LeadUpdate(BaseModel):
    status: Optional[Literal["new", "contacted", "follow_up", "converted", "lost"]] = None
    gender: Optional[str] = None
    dob: Optional[str] = None
    school_institute: Optional[str] = None
    city: Optional[str] = None
    preferred_batch: Optional[str] = None
    source: Optional[str] = None
    counsellor_id: Optional[str] = None
    notes: Optional[str] = None
    next_followup_at: Optional[str] = None

class AttendanceScanRequest(BaseModel):
    student_no: str
    device_signature: Optional[str] = "TER-GATE-01"

class AttendanceOverrideRequest(BaseModel):
    student_id: str
    status: Literal["present", "late"]

# ====== FACTORY MODULE INFRASTRUCTURE ======
def build_erp_router(db, get_current_user, hash_password, verify_password, require_admin):

    # Hard isolation: Dropping implicit parent route dependencies prevents public endpoint pollution
    erp = APIRouter(prefix="/erp", tags=["erp"], dependencies=[])

    # ---- Role guards
    async def require_erp(user: dict = Depends(get_current_user)) -> dict:
        role = user.get("role")
        if role == "admin":
            user["role"] = "super_admin"
            role = "super_admin"
        if role not in ROLES_ALL:
            raise HTTPException(403, "Access Denied: ERP authorization permissions required.")
        return user

    async def require_super(user: dict = Depends(require_erp)) -> dict:
        if user["role"] != "super_admin":
            raise HTTPException(403, "Access Denied: Super admin clearance required.")
        return user

    async def require_manager_plus(user: dict = Depends(require_erp)) -> dict:
        if user["role"] not in {"super_admin", "center_manager"}:
            raise HTTPException(403, "Access Denied: Manager or executive clearance required.")
        return user

    def can_view_branch(user: dict, branch_id: str) -> bool:
        if user["role"] == "super_admin":
            return True
        return user.get("branch_id") == branch_id

    def scope_branch_filter(user: dict, branch_id_param: Optional[str] = None) -> dict:
        if user["role"] == "super_admin":
            return {"branch_id": branch_id_param} if branch_id_param else {}
        if not user.get("branch_id"):
            raise HTTPException(403, "Context Error: User profile has no active branch assignment node.")
        if branch_id_param and branch_id_param != user["branch_id"]:
            raise HTTPException(403, "Access Denied: Cross-branch query parameter operations rejected.")
        return {"branch_id": user["branch_id"]}

    # ---- Audit logger
    async def audit(user: dict, action: str, entity: str, entity_id: str, branch_id: Optional[str] = None, payload: Optional[dict] = None):
        try:
            await db.erp_audit.insert_one({
                "id": new_id(),
                "actor_id": user["id"],
                "actor_email": user.get("email"),
                "actor_role": user.get("role"),
                "action": action,
                "entity": entity,
                "entity_id": entity_id,
                "branch_id": branch_id,
                "payload": payload or {},
                "created_at": now_iso(),
            })
        except Exception:
            pass

    async def broadcast_attendance_event(branch_id: str, event_payload: dict):
        if branch_id in branch_broadcast_queues:
            disconnected_queues = []
            for q in branch_broadcast_queues[branch_id]:
                try:
                    await q.put(event_payload)
                except Exception:
                    disconnected_queues.append(q)
            for dq in disconnected_queues:
                branch_broadcast_queues[branch_id].remove(dq)

    async def gen_receipt_no(branch_id: str) -> str:
        b = await db.centers.find_one({"id": branch_id}, {"_id": 0})
        prefix = "NES"
        if b and b.get("name"):
            prefix = "NES-" + b["name"][:3].upper()
        result = await db.erp_counters.find_one_and_update(
            {"_id": f"receipt_{branch_id}"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True,
        )
        seq = (result or {}).get("seq", 1)
        ymd = datetime.now(timezone.utc).strftime("%y%m")
        return f"{prefix}/{ymd}/{seq:05d}"

    async def gen_student_no(branch_id: str) -> str:
        b = await db.centers.find_one({"id": branch_id}, {"_id": 0})
        prefix = "NES"
        if b and b.get("name"):
            prefix = "NES-" + b["name"][:3].upper()
        result = await db.erp_counters.find_one_and_update(
            {"_id": f"student_{branch_id}"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True,
        )
        seq = (result or {}).get("seq", 1)
        return f"{prefix}-{seq:04d}"

    # ===== ME =====
    @erp.get("/me")
    async def erp_me(user: dict = Depends(require_erp)):
        branch = None
        if user.get("branch_id"):
            branch = await db.centers.find_one({"id": user["branch_id"]}, {"_id": 0})
        return {**user, "branch": branch}

    # ===== AUTOMATED QR ATTENDANCE LOGIC MODULES =====
    @erp.post("/erpattendance/scan")
    async def handle_attendance_scan(payload: AttendanceScanRequest, user: dict = Depends(require_erp)):
        """Parses active raw structural card scanner token text validations asynchronously."""
        current_time = datetime.now(timezone.utc)
        
        student = await db.erp_students.find_one({"student_no": payload.student_no, "status": "active"})
        if not student:
            raise HTTPException(404, "Invalid registration card signature code or suspended profile record match line")
            
        if not can_view_branch(user, student["branch_id"]):
            raise HTTPException(403, "Terminal hardware mapping authorization scope fault")

        min_bound_time = (current_time - timedelta(hours=12)).isoformat()
        double_check = await db.erp_attendance.find_one({
            "student_id": student["id"],
            "scanned_at": {"$gte": min_bound_time}
        })
        if double_check:
            raise HTTPException(422, "Student credential matrix entry sequence has already logged verification for this block loop context")

        calculated_status = "present"
        target_start_hour = 9 
        grace_period_threshold_minutes = 15
        
        local_adjusted_time = current_time + timedelta(hours=5, minutes=30) 
        gate_opening_time = local_adjusted_time.replace(hour=target_start_hour, minute=0, second=0, microsecond=0)
        
        minutes_deviation = (local_adjusted_time - gate_opening_time).total_seconds() / 60.0
        if minutes_deviation > grace_period_threshold_minutes:
            calculated_status = "late"

        log_entry = {
            "id": new_id(),
            "student_id": student["id"],
            "student_no": student["student_no"],
            "full_name": student["full_name"],
            "batch": student.get("batch", "GENERAL COHORT"),
            "branch_id": student["branch_id"],
            "status": calculated_status,
            "mode": "QR Badge Scan via Gate Terminal",
            "device_signature": payload.device_signature,
            "scanned_at": now_iso()
        }
        
        await db.erp_attendance.insert_one(log_entry)
        log_entry.pop("_id", None)

        asyncio.create_task(broadcast_attendance_event(student["branch_id"], log_entry))
        asyncio.create_task(audit(user, "scan_verification", "attendance", log_entry["id"], student["branch_id"], {"status": calculated_status}))

        return log_entry

    @erp.post("/erpattendance/override")
    async def handle_manual_override(payload: AttendanceOverrideRequest, user: dict = Depends(require_manager_plus)):
        """Injects artificial administrative records cleanly bypassing physical scanners."""
        student = await db.erp_students.find_one({"id": payload.student_id})
        if not student:
            raise HTTPException(404, "Target educational tracking record node index empty")
            
        if not can_view_branch(user, student["branch_id"]):
            raise HTTPException(403, "Cross-branch asset operational violation tracking logs intercept")

        log_entry = {
            "id": new_id(),
            "student_id": student["id"],
            "student_no": student["student_no"],
            "full_name": student["full_name"],
            "batch": student.get("batch", "GENERAL COHORT"),
            "branch_id": student["branch_id"],
            "status": payload.status,
            "mode": f"Manual Entry Override by {user.get('name', 'Admin')}",
            "device_signature": "CONSOLE_OVERRIDE_DESK",
            "scanned_at": now_iso()
        }

        await db.erp_attendance.insert_one(log_entry)
        log_entry.pop("_id", None)

        asyncio.create_task(broadcast_attendance_event(student["branch_id"], log_entry))
        asyncio.create_task(audit(user, "manual_override", "attendance", log_entry["id"], student["branch_id"]))
        
        return log_entry

    @erp.get("/erpattendance")
    async def list_attendance_logs(branch_id: Optional[str] = None, user: dict = Depends(require_erp)):
        f = scope_branch_filter(user, branch_id)
        items = await db.erp_attendance.find(f, {"_id": 0}).sort("scanned_at", -1).to_list(1000)
        return items

    @erp.get("/erpattendance/stream/{branch_id}")
    async def live_attendance_sse_stream(branch_id: str, user: dict = Depends(require_erp)):
        if not can_view_branch(user, branch_id):
            raise HTTPException(403, "Stream intercept mapping rejection access token parameter error")

        async def event_generator_loop():
            client_queue = asyncio.Queue()
            branch_broadcast_queues.setdefault(branch_id, []).append(client_queue)
            
            try:
                yield "retry: 10000\ndata: {\"system_status\": \"CONNECTED_STREAM_SYNC_OK\"}\n\n"
                while True:
                    incoming_scan_event = await client_queue.get()
                    yield f"event: attendance_scanned_event\ndata: {json.dumps(incoming_scan_event)}\n\n"
            except asyncio.CancelledError:
                pass
            finally:
                if branch_id in branch_broadcast_queues and client_queue in branch_broadcast_queues[branch_id]:
                    branch_broadcast_queues[branch_id].remove(client_queue)

        return StreamingResponse(event_generator_loop(), media_type="text/event-stream")

    # ===== BRANCHES =====
    @erp.get("/branches")
    async def list_branches(user: dict = Depends(require_erp)):
        if user["role"] == "super_admin":
            items = await db.centers.find({}, {"_id": 0}).to_list(200)
        else:
            if not user.get("branch_id"):
                return []
            items = await db.centers.find({"id": user["branch_id"]}, {"_id": 0}).to_list(10)
        return items

    @erp.patch("/branches/{branch_id}")
    async def update_branch(branch_id: str, payload: BranchUpdate, user: dict = Depends(require_super)):
        patch = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
        if not patch:
            raise HTTPException(400, "Nothing to update")
        await db.centers.update_one({"id": branch_id}, {"$set": patch})
        await audit(user, "update", "branch", branch_id, branch_id, patch)
        return await db.centers.find_one({"id": branch_id}, {"_id": 0})

    # ===== STAFF =====
    @erp.get("/staff")
    async def list_staff(branch_id: Optional[str] = None, user: dict = Depends(require_manager_plus)):
        f: dict = {"role": {"$in": list(ROLES_BRANCH)}}
        if user["role"] == "super_admin":
            if branch_id:
                f["branch_id"] = branch_id
        else:
            f["branch_id"] = user["branch_id"]
        items = await db.users.find(f, {"_id": 0, "password_hash": 0}).to_list(500)
        return items

    @erp.post("/staff")
    async def create_staff(payload: StaffCreate, user: dict = Depends(require_manager_plus)):
        if user["role"] != "super_admin" and payload.branch_id != user.get("branch_id"):
            raise HTTPException(403, "Cannot create staff in another branch")
        if user["role"] == "center_manager" and payload.role == "center_manager":
            raise HTTPException(403, "Only super admin can promote a center manager")
        if await db.users.find_one({"email": payload.email.lower()}):
            raise HTTPException(409, "Email already exists")
        if not await db.centers.find_one({"id": payload.branch_id}):
            raise HTTPException(400, "Branch not found")
        doc = {
            "id": new_id(),
            "name": payload.name,
            "email": payload.email.lower(),
            "phone": payload.phone,
            "role": payload.role,
            "branch_id": payload.branch_id,
            "password_hash": hash_password(payload.password),
            "active": True,
            "created_at": now_iso(),
        }
        await db.users.insert_one(doc)
        doc.pop("_id", None); doc.pop("password_hash", None)
        if payload.role == "center_manager":
            await db.centers.update_one({"id": payload.branch_id}, {"$set": {"manager_user_id": doc["id"]}})
        await audit(user, "create", "staff", doc["id"], payload.branch_id, {"role": payload.role})
        return doc

    @erp.patch("/staff/{staff_id}")
    async def update_staff(staff_id: str, payload: StaffUpdate, user: dict = Depends(require_manager_plus)):
        target = await db.users.find_one({"id": staff_id}, {"_id": 0, "password_hash": 0})
        if not target:
            raise HTTPException(404, "Staff not found")
        if user["role"] != "super_admin" and target.get("branch_id") != user.get("branch_id"):
            raise HTTPException(403, "Cross-branch staff edit denied")
        patch = payload.dict(exclude_unset=True)
        new_pwd = patch.pop("new_password", None)
        if new_pwd:
            patch["password_hash"] = hash_password(new_pwd)
        if "role" in patch and user["role"] == "center_manager" and patch["role"] == "center_manager":
            raise HTTPException(403, "Only super admin can promote to manager")
        if patch:
            await db.users.update_one({"id": staff_id}, {"$set": patch})
        await audit(user, "update", "staff", staff_id, target.get("branch_id"), {"fields": list(patch.keys())})
        return await db.users.find_one({"id": staff_id}, {"_id": 0, "password_hash": 0})

    @erp.delete("/staff/{staff_id}")
    async def deactivate_staff(staff_id: str, user: dict = Depends(require_manager_plus)):
        target = await db.users.find_one({"id": staff_id}, {"_id": 0})
        if not target:
            raise HTTPException(404, "Staff not found")
        if user["role"] != "super_admin" and target.get("branch_id") != user.get("branch_id"):
            raise HTTPException(403, "Cross-branch denied")
        await db.users.update_one({"id": staff_id}, {"$set": {"active": False}})
        await audit(user, "deactivate", "staff", staff_id, target.get("branch_id"))
        return {"ok": True}

    # ===== STUDENTS =====
    @erp.get("/students")
    async def list_students(
        branch_id: Optional[str] = None,
        q: Optional[str] = None,
        course_id: Optional[str] = None,
        counsellor_id: Optional[str] = None,
        user: dict = Depends(require_erp),
    ):
        f = scope_branch_filter(user, branch_id)
        if course_id:
            f["course_id"] = course_id
        if counsellor_id:
            f["counsellor_id"] = counsellor_id
        if user["role"] == "counsellor":
            f["counsellor_id"] = user["id"]
        if q:
            f["$or"] = [
                {"full_name": {"$regex": q, "$options": "i"}},
                {"contact_phone": {"$regex": q, "$options": "i"}},
                {"student_no": {"$regex": q, "$options": "i"}},
            ]
        items = await db.erp_students.find(f, {"_id": 0}).sort("created_at", -1).to_list(500)
        return items

    @erp.post("/students")
    async def create_student(payload: StudentCreate, user: dict = Depends(require_erp)):
        if user["role"] == "counsellor":
            raise HTTPException(403, "Counsellors cannot create student records directly")
        if not can_view_branch(user, payload.branch_id):
            raise HTTPException(403, "Cross-branch denied")
        if not await db.centers.find_one({"id": payload.branch_id}):
            raise HTTPException(400, "Branch not found")
        if not await db.courses.find_one({"id": payload.course_id}):
            raise HTTPException(400, "Course not found")
        student_no = await gen_student_no(payload.branch_id)
        doc = payload.dict()
        doc.update({
            "id": new_id(),
            "student_no": student_no,
            "status": "active",
            "admission_date": payload.admission_date or now_iso()[:10],
            "created_at": now_iso(),
            "created_by": user["id"],
        })
        await db.erp_students.insert_one(doc)
        doc.pop("_id", None)
        await audit(user, "create", "student", doc["id"], payload.branch_id, {"student_no": student_no})
        return doc

    @erp.get("/students/{student_id}")
    async def get_student(student_id: str, user: dict = Depends(require_erp)):
        s = await db.erp_students.find_one({"id": student_id}, {"_id": 0})
        if not s:
            raise HTTPException(404, "Student not found")
        if not can_view_branch(user, s["branch_id"]):
            raise HTTPException(403, "Cross-branch denied")
        if user["role"] == "counsellor" and s.get("counsellor_id") != user["id"]:
            raise HTTPException(403, "Not your student")
        return s

    @erp.patch("/students/{student_id}")
    async def update_student(student_id: str, payload: StudentUpdate, user: dict = Depends(require_erp)):
        s = await db.erp_students.find_one({"id": student_id}, {"_id": 0})
        if not s:
            raise HTTPException(404, "Student not found")
        if user["role"] == "counsellor":
            raise HTTPException(403, "Counsellors cannot edit student records")
        if not can_view_branch(user, s["branch_id"]):
            raise HTTPException(403, "Cross-branch denied")
        patch = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
        if patch:
            await db.erp_students.update_one({"id": student_id}, {"$set": patch})
        await audit(user, "update", "student", student_id, s["branch_id"], {"fields": list(patch.keys())})
        return await db.erp_students.find_one({"id": student_id}, {"_id": 0})

    # ===== PAYMENTS / RECEIPTS =====
    @erp.post("/payments")
    async def create_payment(payload: PaymentCreate, user: dict = Depends(require_erp)):
        if user["role"] == "counsellor":
            raise HTTPException(403, "Counsellors cannot record payments")
        s = await db.erp_students.find_one({"id": payload.student_id}, {"_id": 0})
        if not s:
            raise HTTPException(404, "Student not found")
        if not can_view_branch(user, s["branch_id"]):
            raise HTTPException(403, "Cross-branch denied")
        receipt_no = await gen_receipt_no(s["branch_id"])
        amount = float(payload.amount)
        if amount <= 0:
            raise HTTPException(400, "Amount must be positive")
        if payload.apply_gst:
            base = round(amount / (1 + (CGST_RATE + SGST_RATE) / 100), 2)
            cgst = round(base * CGST_RATE / 100, 2)
            sgst = round(amount - base - cgst, 2)
        else:
            base = amount
            cgst = 0.0
            sgst = 0.0
        doc = {
            "id": new_id(),
            "receipt_no": receipt_no,
            "student_id": payload.student_id,
            "student_no": s["student_no"],
            "branch_id": s["branch_id"],
            "course_id": s["course_id"],
            "amount": amount,
            "base_amount": base,
            "cgst": cgst,
            "sgst": sgst,
            "cgst_rate": CGST_RATE if payload.apply_gst else 0,
            "sgst_rate": SGST_RATE if payload.apply_gst else 0,
            "mode": payload.mode,
            "next_due_date": payload.next_due_date,
            "transaction_ref": payload.transaction_ref,
            "notes": payload.notes,
            "collected_by": user["id"],
            "collected_by_name": user.get("name"),
            "paid_at": now_iso(),
        }
        await db.erp_payments.insert_one(doc); doc.pop("_id", None)
        await audit(user, "create", "payment", doc["id"], s["branch_id"], {"amount": amount, "receipt_no": receipt_no})
        return doc

    @erp.get("/payments")
    async def list_payments(
        student_id: Optional[str] = None,
        branch_id: Optional[str] = None,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
        user: dict = Depends(require_erp),
    ):
        f = scope_branch_filter(user, branch_id)
        if student_id:
            f["student_id"] = student_id
        if from_date or to_date:
            range_f = {}
            if from_date:
                range_f["$gte"] = from_date
            if to_date:
                range_f["$lte"] = to_date + "T23:59:59"
            f["paid_at"] = range_f
        items = await db.erp_payments.find(f, {"_id": 0}).sort("paid_at", -1).to_list(1000)
        return items

    @erp.get("/students/{student_id}/statement")
    async def student_statement(student_id: str, user: dict = Depends(require_erp)):
        s = await db.erp_students.find_one({"id": student_id}, {"_id": 0})
        if not s:
            raise HTTPException(404, "Student not found")
        if not can_view_branch(user, s["branch_id"]):
            raise HTTPException(403, "Cross-branch denied")
        if user["role"] == "counsellor" and s.get("counsellor_id") != user["id"]:
            raise HTTPException(403, "Not your student")
        payments = await db.erp_payments.find({"student_id": student_id}, {"_id": 0}).sort("paid_at", 1).to_list(1000)
        total_paid = sum(p["amount"] for p in payments)
        scholarship_amt = float(s["total_fee"]) * float(s.get("scholarship_percent", 0)) / 100.0
        discount = float(s.get("discount", 0))
        net_fee = max(float(s["total_fee"]) - scholarship_amt - discount, 0)
        return {
            "student": s,
            "total_fee": float(s["total_fee"]),
            "scholarship_percent": float(s.get("scholarship_percent", 0)),
            "scholarship_amount": scholarship_amt,
            "discount": discount,
            "net_fee": net_fee,
            "total_paid": total_paid,
            "pending": max(net_fee - total_paid, 0),
            "payments": payments,
        }

    @erp.get("/payments/{payment_id}/receipt")
    async def download_receipt(payment_id: str, user: dict = Depends(require_erp)):
        p = await db.erp_payments.find_one({"id": payment_id}, {"_id": 0})
        if not p:
            raise HTTPException(404, "Payment not found")
        if not can_view_branch(user, p["branch_id"]):
            raise HTTPException(403, "Cross-branch denied")
        s = await db.erp_students.find_one({"id": p["student_id"]}, {"_id": 0}) or {}
        b = await db.centers.find_one({"id": p["branch_id"]}, {"_id": 0}) or {}
        c = await db.courses.find_one({"id": p.get("course_id")}, {"_id": 0}) or {}
        prev = await db.erp_payments.find(
            {"student_id": p["student_id"], "paid_at": {"$lt": p["paid_at"]}}, {"_id": 0, "amount": 1}
        ).to_list(500)
        prev_paid = sum(x["amount"] for x in prev)
        scholarship_amt = float(s.get("total_fee", 0)) * float(s.get("scholarship_percent", 0)) / 100.0
        net_fee = max(float(s.get("total_fee", 0)) - scholarship_amt - float(s.get("discount", 0)), 0)
        pdf_bytes = fee_receipt_pdf(p, s, b, c.get("title", "—"), prev_paid, net_fee)
        await audit(user, "download", "receipt", p["id"], p["branch_id"])
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="receipt-{p["receipt_no"].replace("/", "-")}.pdf"'},
        )

    # ===== EXPENSES =====
    @erp.post("/expenses")
    async def create_expense(payload: ExpenseCreate, user: dict = Depends(require_erp)):
        if user["role"] not in {"super_admin", "center_manager", "accountant"}:
            raise HTTPException(403, "Not allowed")
        if user["role"] != "super_admin" and payload.branch_id != user.get("branch_id"):
            raise HTTPException(403, "Cross-branch denied")
        if not await db.centers.find_one({"id": payload.branch_id}):
            raise HTTPException(400, "Branch not found")
        auto_approved = user["role"] in {"super_admin", "center_manager"}
        doc = payload.dict()
        doc.update({
            "id": new_id(),
            "status": "approved" if auto_approved else "pending",
            "expense_date": payload.expense_date or now_iso()[:10],
            "recorded_by": user["id"],
            "recorded_by_name": user.get("name"),
            "approved_by": user["id"] if auto_approved else None,
            "approved_at": now_iso() if auto_approved else None,
            "created_at": now_iso(),
        })
        await db.erp_expenses.insert_one(doc); doc.pop("_id", None)
        await audit(user, "create", "expense", doc["id"], payload.branch_id, {"amount": payload.amount, "category": payload.category})
        return doc

    @erp.get("/expenses")
    async def list_expenses(
        branch_id: Optional[str] = None,
        category: Optional[str] = None,
        status: Optional[str] = None,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
        user: dict = Depends(require_erp),
    ):
        if user["role"] == "counsellor":
            raise HTTPException(403, "Not allowed")
        f = scope_branch_filter(user, branch_id)
        if category:
            f["category"] = category
        if status:
            f["status"] = status
        if from_date or to_date:
            range_f = {}
            if from_date:
                range_f["$gte"] = from_date
            if to_date:
                range_f["$lte"] = to_date
            f["expense_date"] = range_f
        items = await db.erp_expenses.find(f, {"_id": 0}).sort("expense_date", -1).to_list(1000)
        return items

    @erp.post("/expenses/{expense_id}/decision")
    async def decide_expense(expense_id: str, payload: ExpenseDecision, user: dict = Depends(require_manager_plus)):
        e = await db.erp_expenses.find_one({"id": expense_id}, {"_id": 0})
        if not e:
            raise HTTPException(404, "Expense not found")
        if user["role"] != "super_admin" and e["branch_id"] != user.get("branch_id"):
            raise HTTPException(403, "Cross-branch denied")
        new_status = "approved" if payload.decision == "approve" else "rejected"
        await db.erp_expenses.update_one({"id": expense_id}, {"$set": {
            "status": new_status,
            "approved_by": user["id"],
            "approved_at": now_iso(),
            "decision_note": payload.note,
        }})
        await audit(user, payload.decision, "expense", expense_id, e["branch_id"])
        return await db.erp_expenses.find_one({"id": expense_id}, {"_id": 0})

    # ===== LEADS =====
    @erp.post("/leads")
    async def create_lead(payload: LeadCreate, user: dict = Depends(require_erp)):
        if user["role"] not in {"super_admin", "center_manager", "counsellor"}:
            raise HTTPException(403, "Not allowed")
        if user["role"] != "super_admin" and payload.branch_id != user.get("branch_id"):
            raise HTTPException(403, "Cross-branch denied")
        cid = payload.counsellor_id or (user["id"] if user["role"] == "counsellor" else None)
        doc = payload.dict()
        doc.update({
            "id": new_id(),
            "status": "new",
            "counsellor_id": cid,
            "created_at": now_iso(),
            "created_by": user["id"],
        })
        await db.erp_leads.insert_one(doc); doc.pop("_id", None)
        await audit(user, "create", "lead", doc["id"], payload.branch_id)
        return doc

    @erp.get("/leads")
    async def list_leads(branch_id: Optional[str] = None, status: Optional[str] = None, user: dict = Depends(require_erp)):
        f = scope_branch_filter(user, branch_id)
        if user["role"] == "counsellor":
            f["counsellor_id"] = user["id"]
        if status:
            f["status"] = status
        items = await db.erp_leads.find(f, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return items

    @erp.patch("/leads/{lead_id}")
    async def update_lead(lead_id: str, payload: LeadUpdate, user: dict = Depends(require_erp)):
        l = await db.erp_leads.find_one({"id": lead_id}, {"_id": 0})
        if not l:
            raise HTTPException(404, "Lead not found")
        if not can_view_branch(user, l["branch_id"]):
            raise HTTPException(403, "Cross-branch denied")
        if user["role"] == "counsellor" and l.get("counsellor_id") != user["id"]:
            raise HTTPException(403, "Not your lead")

        patch = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}

        # Auto-enroll lead to student list when marked as 'converted'
        converted_student = None
        if patch.get("status") == "converted" and l.get("status") != "converted":
            # Check if student already enrolled with same phone at branch
            existing_student = await db.erp_students.find_one({
                "contact_phone": l["phone"],
                "branch_id": l["branch_id"]
            })
            if not existing_student:
                # Find matching course for target_exam or select first course
                target_exam = l.get("target_exam") or "NEET"
                course = await db.courses.find_one({"category": target_exam}) or await db.courses.find_one({})
                course_id = course["id"] if course else "default"
                total_fee = float(course.get("fee", 50000)) if course else 50000.0

                student_no = await gen_student_no(l["branch_id"])
                student_doc = {
                    "id": new_id(),
                    "student_no": student_no,
                    "full_name": l["name"],
                    "gender": l.get("gender"),
                    "dob": l.get("dob"),
                    "school_institute": l.get("school_institute"),
                    "board": l.get("board"),
                    "category": l.get("category"),
                    "contact_phone": l["phone"],
                    "contact_email": l.get("email"),
                    "emergency_phone": l.get("emergency_phone"),
                    "branch_id": l["branch_id"],
                    "course_id": course_id,
                    "batch": l.get("preferred_batch"),
                    "counsellor_id": l.get("counsellor_id") or user["id"],
                    "status": "active",
                    "total_fee": total_fee,
                    "scholarship_percent": 0.0,
                    "discount": 0.0,
                    "admission_date": now_iso()[:10],
                    "notes": f"Auto-converted from lead (ID: {lead_id}). " + (l.get("notes") or ""),
                    "created_at": now_iso(),
                    "created_by": user["id"],
                }
                await db.erp_students.insert_one(student_doc)
                student_doc.pop("_id", None)
                converted_student = student_doc
                patch["converted_student_id"] = student_doc["id"]
                patch["converted_at"] = now_iso()
                await audit(user, "auto_enroll_from_lead", "student", student_doc["id"], l["branch_id"], {"lead_id": lead_id, "student_no": student_no})
            else:
                patch["converted_student_id"] = existing_student["id"]
                patch["converted_at"] = now_iso()

        if patch:
            await db.erp_leads.update_one({"id": lead_id}, {"$set": patch})
        await audit(user, "update", "lead", lead_id, l["branch_id"], patch)

        updated_lead = await db.erp_leads.find_one({"id": lead_id}, {"_id": 0})
        if converted_student:
            updated_lead["converted_student"] = converted_student
        return updated_lead

    # ===== DASHBOARDS =====
    @erp.get("/dashboard/super")
    async def super_dashboard(user: dict = Depends(require_super)):
        branches = await db.centers.find({}, {"_id": 0}).to_list(100)
        rows = []
        total_rev = 0.0
        total_exp = 0.0
        for b in branches:
            payments = await db.erp_payments.find({"branch_id": b["id"]}, {"_id": 0, "amount": 1}).to_list(10000)
            expenses = await db.erp_expenses.find({"branch_id": b["id"], "status": "approved"}, {"_id": 0, "amount": 1}).to_list(10000)
            students = await db.erp_students.count_documents({"branch_id": b["id"], "status": "active"})
            rev = sum(p["amount"] for p in payments)
            exp = sum(e["amount"] for e in expenses)
            total_rev += rev
            total_exp += exp
            rows.append({
                "branch_id": b["id"],
                "branch_name": b.get("name"),
                "city": b.get("city"),
                "revenue": rev,
                "expense": exp,
                "net": rev - exp,
                "students": students,
            })
        all_students = await db.erp_students.find({"status": "active"}, {"_id": 0}).to_list(10000)
        pending_total = 0.0
        for s in all_students:
            paid = await db.erp_payments.find({"student_id": s["id"]}, {"_id": 0, "amount": 1}).to_list(1000)
            scholarship_amt = float(s["total_fee"]) * float(s.get("scholarship_percent", 0)) / 100.0
            net_fee = max(float(s["total_fee"]) - scholarship_amt - float(s.get("discount", 0)), 0)
            pending_total += max(net_fee - sum(p["amount"] for p in paid), 0)
        return {
            "total_revenue": total_rev,
            "total_expense": total_exp,
            "net_income": total_rev - total_exp,
            "total_pending_fees": pending_total,
            "total_students": sum(r["students"] for r in rows),
            "total_branches": len(rows),
            "branches": rows,
        }

    @erp.get("/dashboard/branch/{branch_id}")
    async def branch_dashboard(branch_id: str, user: dict = Depends(require_erp)):
        if not can_view_branch(user, branch_id):
            raise HTTPException(403, "Cross-branch denied")
        b = await db.centers.find_one({"id": branch_id}, {"_id": 0})
        if not b:
            raise HTTPException(404, "Branch not found")
        payments = await db.erp_payments.find({"branch_id": branch_id}, {"_id": 0}).to_list(10000)
        expenses = await db.erp_expenses.find({"branch_id": branch_id, "status": "approved"}, {"_id": 0}).to_list(10000)
        students = await db.erp_students.find({"branch_id": branch_id, "status": "active"}, {"_id": 0}).to_list(10000)
        leads = await db.erp_leads.find({"branch_id": branch_id}, {"_id": 0}).to_list(10000)
        pending_total = 0.0
        for s in students:
            paid = sum(p["amount"] for p in payments if p["student_id"] == s["id"])
            scholarship_amt = float(s["total_fee"]) * float(s.get("scholarship_percent", 0)) / 100.0
            net_fee = max(float(s["total_fee"]) - scholarship_amt - float(s.get("discount", 0)), 0)
            pending_total += max(net_fee - paid, 0)
        cat_split = {}
        for e in expenses:
            cat_split[e["category"]] = cat_split.get(e["category"], 0) + e["amount"]
        cperf: dict = {}
        for s in students:
            cid = s.get("counsellor_id")
            if cid:
                cperf.setdefault(cid, {"converted": 0})
                cperf[cid]["converted"] += 1
        for l in leads:
            cid = l.get("counsellor_id")
            if cid:
                cperf.setdefault(cid, {"converted": 0})
                cperf[cid].setdefault("leads", 0)
                cperf[cid]["leads"] = cperf[cid].get("leads", 0) + 1
        cnames = {}
        if cperf:
            for u in await db.users.find({"id": {"$in": list(cperf.keys())}}, {"_id": 0, "id": 1, "name": 1}).to_list(100):
                cnames[u["id"]] = u["name"]
        counsellor_rows = [{"counsellor_id": k, "name": cnames.get(k, "—"), "leads": v.get("leads", 0), "converted": v.get("converted", 0)} for k, v in cperf.items()]
        return {
            "branch": b,
            "revenue": sum(p["amount"] for p in payments),
            "expense": sum(e["amount"] for e in expenses),
            "pending_fees": pending_total,
            "student_count": len(students),
            "lead_count": len(leads),
            "expense_by_category": cat_split,
            "counsellor_performance": counsellor_rows,
            "recent_payments": sorted(payments, key=lambda x: x.get("paid_at", ""), reverse=True)[:10],
        }

    # ===== EXPORTS =====
    @erp.get("/exports/payments.xlsx")
    async def export_payments_xlsx(branch_id: Optional[str] = None, user: dict = Depends(require_erp)):
        if user["role"] == "counsellor":
            raise HTTPException(403, "Not allowed")
        f = scope_branch_filter(user, branch_id)
        items = await db.erp_payments.find(f, {"_id": 0}).sort("paid_at", -1).to_list(10000)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Payments"
        ws.append(["Receipt", "Date", "Student No", "Branch", "Amount", "CGST", "SGST", "Mode", "Collected By"])
        for p in items:
            ws.append([p["receipt_no"], p["paid_at"][:10], p["student_no"], p["branch_id"], p["amount"], p.get("cgst", 0), p.get("sgst", 0), p["mode"], p.get("collected_by_name", "")])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": 'attachment; filename="payments.xlsx"'})

    @erp.get("/exports/expenses.xlsx")
    async def export_expenses_xlsx(branch_id: Optional[str] = None, user: dict = Depends(require_erp)):
        if user["role"] == "counsellor":
            raise HTTPException(403, "Not allowed")
        f = scope_branch_filter(user, branch_id)
        items = await db.erp_expenses.find(f, {"_id": 0}).sort("expense_date", -1).to_list(10000)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Expenses"
        ws.append(["Date", "Category", "Description", "Vendor", "Amount", "Status", "Recorded By"])
        for e in items:
            ws.append([e["expense_date"], e["category"], e["description"], e.get("vendor", ""), e["amount"], e["status"], e.get("recorded_by_name", "")])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": 'attachment; filename="expenses.xlsx"'})

    @erp.get("/exports/students.xlsx")
    async def export_students_xlsx(branch_id: Optional[str] = None, user: dict = Depends(require_erp)):
        if user["role"] == "counsellor":
            raise HTTPException(403, "Not allowed")
        f = scope_branch_filter(user, branch_id)
        items = await db.erp_students.find(f, {"_id": 0}).sort("created_at", -1).to_list(10000)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Students"
        ws.append(["Student No", "Name", "Phone", "Email", "Branch", "Course", "Batch", "Total Fee", "Scholarship %", "Status", "Admission Date"])
        for s in items:
            ws.append([s["student_no"], s["full_name"], s.get("contact_phone", ""), s.get("contact_email", ""), s["branch_id"], s["course_id"], s.get("batch", ""), s["total_fee"], s.get("scholarship_percent", 0), s["status"], s.get("admission_date", "")])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": 'attachment; filename="students.xlsx"'})

    # ============================================================================
    # COMPREHENSIVE DAILY ATTENDANCE EXCEL EXPORTER ENDPOINT
    # ============================================================================
    @erp.get("/erpattendance/exports/attendance_today.xlsx")
    async def export_todays_attendance_matrix(branch_id: Optional[str] = None, user: dict = Depends(require_erp)):
        if user["role"] == "counsellor":
            raise HTTPException(403, "Access Denied: Administrative permission clearance required.")
            
        f = scope_branch_filter(user, branch_id)
        
        # Calculate localized ISO boundaries tracking today's date metrics strictly
        today_start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        f["scanned_at"] = {"$gte": f"{today_start_date}T00:00:00"}
        
        # Pull raw real-time gate log entries matching system configurations
        raw_attendance_logs = await db.erp_attendance.find(f).sort("scanned_at", 1).to_list(10000)
        
        # Extract metadata mapping configurations safely via isolated asynchronous db list blocks
        courses_cursor = db.courses.find({}, {"id": 1, "title": 1})
        all_courses = {c["id"]: c.get("title", "Unknown Track") for c in await courses_cursor.to_list(1000)}
        
        students_cursor = db.erp_students.find({}, {"id": 1, "course_id": 1, "contact_phone": 1, "parent_phone": 1})
        all_students = {s["id"]: s for s in await students_cursor.to_list(10000)}

        # Construct structural memory grid mapped under [Course_Title][Batch_Tag] matrices
        grouped_workbook_data = {}
        for entry in raw_attendance_logs:
            student_meta = all_students.get(entry["student_id"], {})
            course_id = student_meta.get("course_id", "GENERAL")
            course_title = all_courses.get(course_id, "General Roster").replace("/", "-")[:30] # Spreadsheet sheetname safety truncation cap
            
            raw_batch = entry.get("batch")
            batch_tag = "UNASSIGNED_BATCH" if not raw_batch else str(raw_batch).replace("/", "-").upper()
            
            grouped_workbook_data.setdefault(course_title, {}).setdefault(batch_tag, []).append(entry)

        # Initialize raw workbook compiler container
        wb = openpyxl.Workbook()
        
        if not grouped_workbook_data:
            ws = wb.active
            ws.title = "No Attendance Today"
            ws.append(["System Status Note", "No entry gate verification logs recorded yet for today."])
        else:
            for class_title, batches_sub_dict in grouped_workbook_data.items():
                ws = wb.create_sheet(title=class_title)
                
                for batch_code, record_rows in batches_sub_dict.items():
                    # Format sheet data alignment blocks
                    ws.append([]) 
                    ws.append([f"CLASS TRACK: {class_title} — BATCH COHORT: {batch_code}"])
                    ws.append([
                        "Enrollment No", 
                        "Student Profile Full Name", 
                        "Student Mobile No", 
                        "Parent Mobile No", 
                        "Check-In Verified Clock", 
                        "Status Block Status", 
                        "Gate Gateway Mode"
                    ])
                    
                    for r in record_rows:
                        # Re-verify tracking loops context records cleanly from mapping cache
                        meta = all_students.get(r["student_id"], {})
                        student_phone = meta.get("contact_phone", "—")
                        parent_phone = meta.get("parent_phone", "—")
                        
                        scan_time = r.get("scanned_at", "")
                        local_time_string = scan_time[11:16] if len(scan_time) > 16 else scan_time
                        
                        ws.append([
                            r.get("student_no", "—"), 
                            r.get("full_name", "—"), 
                            student_phone,
                            parent_phone,
                            local_time_string, 
                            str(r.get("status", "PRESENT")).upper(), 
                            r.get("mode", "Scan Entry")
                        ])
            
            # Flush out structural artifact default workspace sheets safely
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb["Sheet"])
            elif "Sheet" in wb.sheetnames:
                wb["Sheet"].title = "Empty Roster Summary"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        asyncio.create_task(audit(user, "export_todays_metrics", "attendance", "today_xlsx", branch_id))
        
        return StreamingResponse(
            buf, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="attendance_report_{today_start_date}.xlsx"'}
        )

    # ===== AUDIT LOG (super_admin) =====
    @erp.get("/audit")
    async def audit_log(branch_id: Optional[str] = None, limit: int = 200, user: dict = Depends(require_super)):
        f = {"branch_id": branch_id} if branch_id else {}
        items = await db.erp_audit.find(f, {"_id": 0}).sort("created_at", -1).to_list(min(limit, 1000))
        return items

    # ===== META =====
    @erp.get("/meta")
    async def meta(user: dict = Depends(require_erp)):
        return {
            "expense_categories": EXPENSE_CATEGORIES,
            "payment_modes": PAYMENT_MODES,
            "lead_statuses": LEAD_STATUSES,
            "cgst_rate": CGST_RATE,
            "sgst_rate": SGST_RATE,
            "roles": list(ROLES_ALL),
        }

    return erp


# ====== ERP seed (idempotent): ensures indexes ======
async def erp_seed(db, hash_password):
    await db.erp_students.create_index("student_no", unique=True, sparse=True)
    await db.erp_students.create_index([("branch_id", 1), ("created_at", -1)])
    await db.erp_payments.create_index("receipt_no", unique=True, sparse=True)
    await db.erp_payments.create_index([("branch_id", 1), ("paid_at", -1)])
    await db.erp_payments.create_index([("student_id", 1), ("paid_at", 1)])
    await db.erp_expenses.create_index([("branch_id", 1), ("expense_date", -1)])
    await db.erp_leads.create_index([("branch_id", 1), ("status", 1)])
    await db.erp_audit.create_index([("created_at", -1)])
    await db.erp_attendance.create_index([("branch_id", 1), ("scanned_at", -1)])
    await db.erp_attendance.create_index([("student_id", 1), ("scanned_at", -1)])