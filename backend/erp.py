import io
import asyncio
from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, EventSourceResponse
import openpyxl

erp = APIRouter()

# Helper placeholder for system audit logger and scope filtering
async def audit(user, action, resource, target, branch_id): pass
def scope_branch_filter(user, branch_id): return {"branch_id": branch_id} if branch_id else {}

@erp.get("/erpattendance")
async def list_attendance_logs(branch_id: Optional[str] = None, user: dict = Depends(require_erp)):
    f = scope_branch_filter(user, branch_id)
    return await db.erp_attendance.find(f).sort("scanned_at", -1).to_list(100)

@erp.post("/erpattendance/scan")
async def submit_attendance_scan(body: dict, user: dict = Depends(require_erp)):
    # Simulates verification logic against student hardware tags database
    student_no = body.get("student_no")
    student = await db.erp_students.find_one({"student_no": student_no})
    if not student:
        raise HTTPException(404, "Invalid Student Enrollment Identification Token.")
        
    log_entry = {
        "student_id": student["id"],
        "student_no": student["student_no"],
        "full_name": student["full_name"],
        "branch_id": student["branch_id"],
        "batch": student.get("batch_name", "UNASSIGNED"),
        "status": "present",
        "mode": "Scan Entry",
        "scanned_at": datetime.now(timezone.utc).isoformat()
    }
    await db.erp_attendance.insert_one(log_entry)
    # Broadcaster queues event here for live SSE streams...
    return log_entry

@erp.post("/erpattendance/override")
async def submit_manual_attendance_override(body: dict, user: dict = Depends(require_erp)):
    if user["role"] not in ["super_admin", "admin", "center_manager"]:
        raise HTTPException(403, "Insufficient hierarchy execution permissions.")
    # Implement override write pipeline here
    return {"status": "success"}

@erp.get("/erpattendance/stream/{branch_id}")
async def get_attendance_live_stream(branch_id: str):
    async def event_generator():
        while True:
            # Pushes updates down the wire to connection instances asynchronously
            await asyncio.sleep(10)
    return EventSourceResponse(event_generator())

@erp.get("/erpattendance/exports/attendance_today.xlsx")
async def export_todays_attendance_matrix(branch_id: Optional[str] = None, user: dict = Depends(require_erp)):
    if user["role"] == "counsellor":
        raise HTTPException(403, "Access Denied: Administrative permission clearance required.")
        
    f = scope_branch_filter(user, branch_id)
    today_start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    f["scanned_at"] = {"$gte": f"{today_start_date}T00:00:00"}
    
    raw_attendance_logs = await db.erp_attendance.find(f).sort("scanned_at", 1).to_list(10000)
    
    courses_cursor = db.courses.find({}, {"id": 1, "title": 1})
    all_courses = {c["id"]: c.get("title", "Unknown Track") for c in await courses_cursor.to_list(1000)}
    
    students_cursor = db.erp_students.find({}, {"id": 1, "course_id": 1})
    all_students = {s["id"]: s for s in await students_cursor.to_list(10000)}

    grouped_workbook_data = {}
    for entry in raw_attendance_logs:
        student_meta = all_students.get(entry["student_id"], {})
        course_id = student_meta.get("course_id", "GENERAL")
        course_title = all_courses.get(course_id, "General Roster").replace("/", "-")[:30]
        
        raw_batch = entry.get("batch")
        batch_tag = "UNASSIGNED_BATCH" if not raw_batch else str(raw_batch).replace("/", "-").upper()
        
        grouped_workbook_data.setdefault(course_title, {}).setdefault(batch_tag, []).append(entry)

    wb = openpyxl.Workbook()
    
    if not grouped_workbook_data:
        ws = wb.active
        ws.title = "No Attendance Today"
        ws.append(["System Status Note", "No entry gate verification logs recorded yet for today."])
    else:
        for class_title, batches_sub_dict in grouped_workbook_data.items():
            ws = wb.create_sheet(title=class_title)
            for batch_code, record_rows in batches_sub_dict.items():
                ws.append([]) 
                ws.append([f"BATCH COHORT: {batch_code}"])
                ws.append(["Enrollment No", "Student Profile Full Name", "Check-In Verified Clock", "Status Block Status", "Gate Gateway Mode"])
                
                for r in record_rows:
                    scan_time = r.get("scanned_at", "")
                    local_time_string = scan_time[11:16] if len(scan_time) > 16 else scan_time
                    ws.append([
                        r.get("student_no", "—"), r.get("full_name", "—"), 
                        local_time_string, str(r.get("status", "PRESENT")).upper(), r.get("mode", "Scan Entry")
                    ])
        
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
        elif "Sheet" in wb.sheetnames:
            wb["Sheet"].title = "Empty Roster Summary"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    asyncio.create_task(audit(user, "export_todays_metrics", "attendance", "today_xlsx", branch_id))
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="attendance_report_{today_start_date}.xlsx"'}
    )