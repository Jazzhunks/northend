# 1. LOAD ENVIRONMENT VARIABLES FIRST
import os
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# NVIDIA API Configuration dynamically loaded from environment
NVIDIA_BASE_URL = os.getenv("NVIDIA_BASE_URL")
NVIDIA_API_KEY = os.getenv("NVIDIA_API_KEY")
NVIDIA_MODEL = os.getenv("NVIDIA_MODEL", "nvidia/nemotron-3.5-lightning-30b-a3b")

# 2. STANDARD LIBRARY IMPORTS
import io
import re
import uuid
import random
import logging
import asyncio
import inspect
import tempfile
import shutil
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Dict, Any

# 3. EXTERNAL DEPENDENCIES
import bcrypt
import jwt
import openpyxl
from fastapi import FastAPI, APIRouter, BackgroundTasks, HTTPException, UploadFile, File, Depends, Request, Response, Query, Form
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

# 4. INTERNAL CLIENTS (Loaded after environment is populated)
from storage_client import init_storage, put_object, get_object, aclose as storage_aclose, APP_NAME
from email_client import (
    email_enrollment_received, email_scholarship_received,
    email_job_app_received, email_admin_notification,
    email_scholarship_result_published,
)
from pdf_client import admit_card_pdf, result_card_pdf
from whatsapp_client import send_whatsapp_admit_card, send_whatsapp_wath_carnival
from whatsapp_notifier import broadcast_scholarship_details


# ---------- Constants & Helpers ----------
ALLOWED_UPLOAD_TYPES = {
    "application/pdf": "pdf",
    "image/jpeg": "jpg", "image/jpg": "jpg",
    "image/png": "png", "image/webp": "webp",
    "video/mp4": "mp4", "video/webm": "webm", "video/quicktime": "mov",
}
MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 MB

# ---------- Rate Limiter & Security Lockout ----------
_rate_limit_store: dict[str, list[float]] = {}
_failed_login_store: dict[str, list[float]] = {}
_lockout_until: dict[str, float] = {}

def check_rate_limit(key: str, max_requests: int = 30, window_seconds: int = 60):
    now = datetime.now(timezone.utc).timestamp()
    timestamps = [t for t in _rate_limit_store.get(key, []) if now - t < window_seconds]
    if len(timestamps) >= max_requests:
        raise HTTPException(status_code=429, detail="Too many requests. Please try again later.")
    timestamps.append(now)
    _rate_limit_store[key] = timestamps

def check_login_lockout(key: str):
    now = datetime.now(timezone.utc).timestamp()
    until = _lockout_until.get(key, 0)
    if now < until:
        remaining = int(until - now)
        raise HTTPException(
            status_code=429,
            detail=f"Account temporarily locked due to repeated failed login attempts. Try again in {remaining} seconds."
        )

def record_failed_login(key: str, max_failures: int = 5, window_seconds: int = 900, lockout_seconds: int = 900):
    now = datetime.now(timezone.utc).timestamp()
    failures = [t for t in _failed_login_store.get(key, []) if now - t < window_seconds]
    failures.append(now)
    _failed_login_store[key] = failures
    if len(failures) >= max_failures:
        _lockout_until[key] = now + lockout_seconds
        _failed_login_store.pop(key, None)

def reset_login_failures(key: str):
    _failed_login_store.pop(key, None)
    _lockout_until.pop(key, None)

def _sanitize_venue(venue: Optional[str]) -> str:
    """Sanitize standard venues or preserve custom school venue names. Merges Srinagar/90FT"""
    if not venue:
        return "90 FT"
    v_clean = venue.strip()
    v_lower = v_clean.lower()
    
    # Merge variations of 90 FT and Srinagar
    if v_lower in ("90 ft", "90ft", "srinagar"):
        return "90 FT"
        
    for allowed in ["Anantnag", "Sopore", "Zakura", "Parraypora"]:
        if allowed.lower() == v_lower:
            return allowed
            
    return v_clean

async def _run_maybe_async(func, *args, **kwargs):
    """Run either an async or synchronous client function safely."""
    if inspect.iscoroutinefunction(func):
        return await func(*args, **kwargs)
    return await asyncio.to_thread(func, *args, **kwargs)


def _safe_send_whatsapp_admit_card(*args, **kwargs) -> None:
    """Sync wrapper for FastAPI BackgroundTasks for standard admit cards."""
    try:
        asyncio.run(_run_maybe_async(send_whatsapp_admit_card, *args, **kwargs))
    except Exception as e:
        logging.error(f"Background WhatsApp task failed: {e}")

def _safe_send_whatsapp_wath_carnival(*args, **kwargs) -> None:
    """Sync wrapper for FastAPI BackgroundTasks for WATH Carnival messages."""
    try:
        asyncio.run(_run_maybe_async(send_whatsapp_wath_carnival, *args, **kwargs))
    except Exception as e:
        logging.error(f"Background WATH Carnival WhatsApp task failed: {e}")

def export_excel(rows: list, sheet_name: str, filename: str):
    """Helper utility for generating Excel downloads."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([str(r.get(h, "")) for h in headers])
    else:
        ws.append(["No data"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ---------- Setup ----------
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
use_mock = os.environ.get('USE_MOCK_MONGO', '').strip().lower() in ('1', 'true', 'yes')

if use_mock:
    from mongomock_motor import AsyncMongoMockClient
    client = AsyncMongoMockClient()
    db = client[os.environ.get('DB_NAME', 'northend_db')]
else:
    try:
        client = AsyncIOMotorClient(mongo_url)
        try:
            db = client.get_default_database()
            if db is None:
                raise ValueError("no default database in MONGO_URL")
        except Exception:
            db = client[os.environ.get('DB_NAME', 'northend_db')]
    except Exception:
        from mongomock_motor import AsyncMongoMockClient
        client = AsyncMongoMockClient()
        db = client[os.environ.get('DB_NAME', 'northend_db')]

app = FastAPI(title="Unacademy Offline Centre API")
api = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"

def jwt_secret() -> str:
    secret = os.getenv("JWT_SECRET")
    if not secret:
        raise RuntimeError("JWT_SECRET is not configured on the server")
    return secret

def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False

def create_access_token(uid: str, email: str, role: str) -> str:
    return jwt.encode(
        {"sub": uid, "email": email, "role": role,
         "exp": datetime.now(timezone.utc) + timedelta(minutes=60),
         "type": "access"},
        jwt_secret(), algorithm=JWT_ALGORITHM
    )

def create_refresh_token(uid: str) -> str:
    return jwt.encode(
        {"sub": uid, "exp": datetime.now(timezone.utc) + timedelta(days=7),
         "type": "refresh"},
        jwt_secret(), algorithm=JWT_ALGORITHM
    )

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=True, samesite="lax", max_age=3600, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True, samesite="lax", max_age=604800, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        ah = request.headers.get("Authorization", "")
        if ah.startswith("Bearer "):
            token = ah[7:]
    
    if not token:
        token = request.query_params.get("token")

    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(401, "Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(401, "User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin access required")
    return user

async def require_super_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") not in ("admin", "super_admin"):
        raise HTTPException(403, "Super admin access required")
    return user

async def require_school(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "school":
        raise HTTPException(403, "School access required")
    return user

# ---------- Models ----------
class RegisterIn(BaseModel):
    name: str
    email: EmailStr
    password: str
    phone: Optional[str] = None
    school_name: Optional[str] = None
    address: Optional[str] = None
    district: Optional[str] = None
    school_type: Optional[Literal["middle", "high", "higher_secondary"]] = None

class LoginIn(BaseModel):
    email: EmailStr
    password: str

ALLOWED_CATEGORIES = ["NEET", "IIT-JEE", "Foundation", "CBSE", "JKBOSE"]

class CourseIn(BaseModel):
    title: str
    category: Literal["NEET", "IIT-JEE", "Foundation", "CBSE", "JKBOSE"]
    duration: str
    fee: int
    description: str
    syllabus: List[str] = []
    faculty: List[str] = []
    features: List[str] = []
    scholarship_available: bool = True
    featured: bool = False
    image_url: Optional[str] = None

class ScholarshipIn(BaseModel):
    title: str
    description: str
    exam_date: str
    deadline: str
    eligibility: str
    venue: Optional[str] = None
    available_venues: List[str] = []
    exam_time: Optional[str] = None
    total_marks: Optional[int] = 100
    whatsapp_community_url: Optional[str] = None
    active: bool = True
    is_featured: bool = False
    kind: Literal["scholarship", "wath"] = "scholarship"
    type: Literal["general", "school"] = "general"

class ScholarshipApplicationIn(BaseModel):
    name: str
    email: EmailStr
    phone: str
    school: str
    standard: str
    target_exam: str
    city: str
    scholarship_id: Optional[str] = None
    venue: Optional[str] = None
    address: Optional[str] = None
    district: Optional[str] = None
    # Extra candidate details (for official admit card)
    father_name: Optional[str] = None
    gender: Optional[str] = None
    dob: Optional[str] = None
    # WATH Carnival fields
    carnival_id: Optional[str] = None
    chosen_date: Optional[str] = None   # YYYY-MM-DD
    chosen_slot_time: Optional[str] = None  # "10:00 AM"

class ScholarshipApplicationUpdateIn(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    school: Optional[str] = None
    standard: Optional[str] = None
    target_exam: Optional[str] = None
    city: Optional[str] = None
    venue: Optional[str] = None
    address: Optional[str] = None
    district: Optional[str] = None
    status: Optional[str] = None

class ScholarshipResultIn(BaseModel):
    marks_obtained: float
    total_marks: float = 100
    rank: Optional[int] = None
    percentile: Optional[float] = None
    scholarship_percentage: int
    remarks: Optional[str] = None
    publish: bool = False

class ScholarshipLookupIn(BaseModel):
    phone: str
    application_no: str

class SchoolVisitIn(BaseModel):
    scholarship_id: str
    preferred_date: str
    preferred_slot_time: str
    notes: Optional[str] = None

class SchoolVisitOut(BaseModel):
    id: str
    school_id: str
    school_name: str
    scholarship_id: str
    preferred_date: str
    preferred_slot_time: str
    status: Literal["pending", "approved", "rejected"] = "pending"
    admin_notes: Optional[str] = None
    created_at: str

class SchoolBulkStudentRow(BaseModel):
    name: str
    mobile: str
    current_class: str
    course: str

class SchoolBulkRegisterResult(BaseModel):
    processed: int
    created: int
    skipped: int
    errors: List[str] = []

class AttendanceMarkIn(BaseModel):
    token: str
    application_no: str
    venue: str
    status: Literal["present", "absent"] = "present"

class EnrollmentIn(BaseModel):
    course_id: str
    name: str
    email: EmailStr
    phone: str
    address: str
    center: str
    id_proof_url: Optional[str] = None

class JobIn(BaseModel):
    title: str
    department: str
    location: str
    type: str
    description: str
    requirements: List[str] = []
    active: bool = True
    is_featured: bool = False

class JobApplicationIn(BaseModel):
    name: str
    email: EmailStr
    phone: str
    job_id: str
    qualification: str
    experience: str
    subject_expertise: Optional[str] = None
    preferred_location: str
    cover_letter: Optional[str] = None
    resume_url: Optional[str] = None

class NoticeIn(BaseModel):
    title: str
    content: str
    category: str = "General"
    pinned: bool = False
    is_featured: bool = False

class ResultIn(BaseModel):
    student_name: str
    exam: str
    rank: str
    year: int
    course: str
    photo_url: Optional[str] = None
    quote: Optional[str] = None

class ContactIn(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    subject: str
    message: str

class CenterIn(BaseModel):
    name: str
    city: str
    address: str
    phone: str
    timing: str = "8:00 AM – 8:00 PM"
    lat: float = 34.0837
    lng: float = 74.7973

class TestimonialIn(BaseModel):
    name: str
    role: str
    quote: str

class GalleryItemIn(BaseModel):
    title: str
    description: Optional[str] = None
    media_type: str = "image"  # image | video | text
    media_url: Optional[str] = None
    category: Optional[str] = None
    order: int = 0

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def new_id():
    return str(uuid.uuid4())

def slugify(text: str) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"[\s_-]+", "-", text).strip("-")
    return text or "item"

async def unique_slug(collection: str, base: str, exclude_id: str | None = None) -> str:
    slug = slugify(base)
    candidate = slug
    i = 2
    while True:
        q: Dict[str, Any] = {"slug": candidate}
        if exclude_id:
            q["id"] = {"$ne": exclude_id}
        if not await db[collection].find_one(q):
            return candidate
        candidate = f"{slug}-{i}"
        i += 1

# ---------- Auth Routes ----------
@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower().strip()
    if len(email) > 254 or len(payload.password) > 128:
        raise HTTPException(400, "Invalid payload length")

    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already registered")
    user_id = new_id()
    role = "school" if payload.school_name else "student"
    doc = {
        "id": user_id, "name": payload.name, "email": email,
        "phone": payload.phone, "role": role,
        "password_hash": hash_password(payload.password),
        "created_at": now_iso(),
    }
    if role == "school":
        doc.update({
            "school_name": payload.school_name,
            "address": payload.address,
            "district": payload.district,
            "school_type": payload.school_type,
        })
    await db.users.insert_one(doc)
    access = create_access_token(user_id, email, role)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    doc.pop("password_hash")
    doc.pop("_id", None)
    return {"user": doc, "access_token": access}

@api.post("/auth/login")
async def login(payload: LoginIn, request: Request, response: Response):
    email = payload.email.lower().strip()
    client_ip = request.client.host if request.client else "unknown"
    lockout_key = f"{email}:{client_ip}"
    check_login_lockout(lockout_key)

    if len(email) > 254 or len(payload.password) > 128:
        record_failed_login(lockout_key)
        raise HTTPException(401, "Invalid email or password")

    user = await db.users.find_one({"email": email})
    dummy_hash = "$2b$12$UnV2ZWRuZWVkTG9naW5IYXJkZW5lZFNlY3VyaXR5R29vZA=="
    target_hash = user["password_hash"] if user else dummy_hash
    password_correct = verify_password(payload.password, target_hash)

    if not user or not password_correct:
        record_failed_login(lockout_key)
        raise HTTPException(401, "Invalid email or password")
        
    reset_login_failures(lockout_key)
    access = create_access_token(user["id"], email, user["role"])
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    user.pop("password_hash")
    user.pop("_id", None)
    return {"user": user, "access_token": access}

@api.post("/auth/logout")
async def logout(response: Response):
    response.set_cookie("access_token", "", httponly=True, secure=True, samesite="none", max_age=0, path="/")
    response.set_cookie("refresh_token", "", httponly=True, secure=True, samesite="none", max_age=0, path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

# ---------- Public stats ----------
@api.get("/stats")
async def stats():
    courses = await db.courses.count_documents({})
    students = await db.users.count_documents({"role": "student"})
    enrollments = await db.enrollments.count_documents({})
    centers = await db.centers.count_documents({})
    selections = await db.results.count_documents({})
    return {
        "students_trained": max(students * 50 + 12000, 12000),
        "selections": max(selections * 8 + 850, 850),
        "educators": 65,
        "centers": max(centers, 6),
        "courses": courses,
        "enrollments": enrollments,
    }

# ---------- Featured highlight ----------
FEATURED_COLLECTIONS = ["notices", "jobs", "scholarships"]

async def _clear_featured_except(keep_coll: str | None, keep_id: str | None):
    for coll in FEATURED_COLLECTIONS:
        q = {"is_featured": True}
        if coll == keep_coll and keep_id:
            q["id"] = {"$ne": keep_id}
        await db[coll].update_many(q, {"$set": {"is_featured": False}})

@api.get("/featured")
async def get_featured():
    for coll, kind in [("notices", "notice"), ("jobs", "job"), ("scholarships", "scholarship")]:
        doc = await db[coll].find_one({"is_featured": True}, {"_id": 0})
        if doc:
            doc["kind"] = kind
            return doc
    return None

@api.post("/admin/feature")
async def set_featured(kind: str = Query(...), item_id: str = Query(...), _admin = Depends(require_admin)):
    if kind == "clear":
        await _clear_featured_except(None, None)
        return {"ok": True, "featured": None}
    coll_map = {"notice": "notices", "job": "jobs", "scholarship": "scholarships"}
    coll = coll_map.get(kind)
    if not coll:
        raise HTTPException(400, "kind must be notice|job|scholarship|clear")
    target = await db[coll].find_one({"id": item_id})
    if not target:
        raise HTTPException(404, f"{kind} not found")
    await _clear_featured_except(coll, item_id)
    await db[coll].update_one({"id": item_id}, {"$set": {"is_featured": True}})
    return {"ok": True, "kind": kind, "id": item_id}

# ---------- Courses ----------
@api.get("/courses")
async def list_courses(category: Optional[str] = None, featured: Optional[bool] = None):
    q = {}
    if category: q["category"] = category
    if featured is not None: q["featured"] = featured
    items = await db.courses.find(q, {"_id": 0}).sort("created_at", -1).to_list(200)
    return items

@api.get("/courses/{cid}")
async def get_course(cid: str):
    c = await db.courses.find_one({"$or": [{"id": cid}, {"slug": cid}]}, {"_id": 0})
    if not c: raise HTTPException(404, "Course not found")
    return c

@api.post("/courses")
async def create_course(payload: CourseIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["slug"] = await unique_slug("courses", doc["title"])
    doc["created_at"] = now_iso()
    await db.courses.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/courses/{cid}")
async def update_course(cid: str, payload: CourseIn, _admin = Depends(require_admin)):
    data = payload.model_dump()
    data["slug"] = await unique_slug("courses", data["title"], exclude_id=cid)
    res = await db.courses.update_one({"id": cid}, {"$set": data})
    if not res.matched_count: raise HTTPException(404, "Course not found")
    return await db.courses.find_one({"id": cid}, {"_id": 0})

@api.delete("/courses/{cid}")
async def delete_course(cid: str, _admin = Depends(require_admin)):
    await db.courses.delete_one({"id": cid})
    return {"ok": True}

# ---------- Scholarships ----------
@api.get("/scholarships")
async def list_scholarships(include_wath: bool = False, type: Optional[str] = Query(None)):
    """Public scholarships list. WATH campaigns are excluded by default so they only appear on /wath."""
    q: Dict[str, Any] = {}
    if not include_wath:
        q["$and"] = [
            {"$or": [{"kind": {"$exists": False}}, {"kind": {"$ne": "wath"}}]},
            {"$or": [{"title": {"$not": {"$regex": "^WATH", "$options": "i"}}}, {"kind": "scholarship"}]},
        ]
    if type:
        q["$or"] = [
            {"type": type},
            {"type": {"$exists": False}},
        ]
    items = await db.scholarships.find(q, {"_id": 0, "examiner_token": 0}).sort("created_at", -1).to_list(100)
    return items

@api.get("/admin/scholarships")
async def list_scholarships_admin(_admin = Depends(require_admin)):
    return await db.scholarships.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)

@api.post("/scholarships")
async def create_scholarship(payload: ScholarshipIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["slug"] = await unique_slug("scholarships", doc["title"])
    doc["created_at"] = now_iso()
    doc["examiner_token"] = uuid.uuid4().hex
    if doc.get("available_venues"):
        doc["available_venues"] = [_sanitize_venue(v) for v in doc["available_venues"]]
    await db.scholarships.insert_one(doc)
    if doc.get("is_featured"):
        await _clear_featured_except("scholarships", doc["id"])
    doc.pop("_id", None)
    return doc

@api.put("/scholarships/{sid}")
async def update_scholarship(sid: str, payload: ScholarshipIn, _admin = Depends(require_admin)):
    data = payload.model_dump()
    data["slug"] = await unique_slug("scholarships", data["title"], exclude_id=sid)
    if data.get("available_venues"):
        data["available_venues"] = [_sanitize_venue(v) for v in data["available_venues"]]
    await db.scholarships.update_one({"id": sid}, {"$set": data})
    if data.get("is_featured"):
        await _clear_featured_except("scholarships", sid)
    return await db.scholarships.find_one({"id": sid}, {"_id": 0})

@api.post("/admin/scholarships/{sid}/regenerate-token")
async def regenerate_examiner_token(sid: str, _admin = Depends(require_admin)):
    new_token = uuid.uuid4().hex
    res = await db.scholarships.update_one({"id": sid}, {"$set": {"examiner_token": new_token}})
    if not res.matched_count:
        raise HTTPException(404, "Campaign not found")
    return {"examiner_token": new_token}

@api.delete("/scholarships/{sid}")
async def delete_scholarship(sid: str, _admin = Depends(require_admin)):
    await db.scholarships.delete_one({"id": sid})
    return {"ok": True}

@api.post("/scholarship-applications")
async def apply_scholarship(payload: ScholarshipApplicationIn, background: BackgroundTasks, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    check_rate_limit(f"scholarship_app:{client_ip}", max_requests=15, window_seconds=60)

    # -------- Resolve target: scholarship campaign OR WATH Carnival --------
    campaign: Optional[Dict[str, Any]] = None
    carnival: Optional[Dict[str, Any]] = None
    campaign_kind = "scholarship"

    if payload.carnival_id:
        carnival = await db.wath_carnivals.find_one({"id": payload.carnival_id}, {"_id": 0})
        if not carnival or not carnival.get("active", True):
            raise HTTPException(404, "Carnival not found or inactive")
        if not payload.chosen_date or not payload.chosen_slot_time:
            raise HTTPException(400, "Please pick an exam date and time slot")
        campaign_kind = "carnival"
    elif payload.scholarship_id:
        campaign = await db.scholarships.find_one({"id": payload.scholarship_id}, {"_id": 0})
        if not campaign:
            raise HTTPException(404, "Scholarship campaign not found")
        if not campaign.get("active"):
            raise HTTPException(400, "Scholarship campaign is closed")
        campaign_kind = campaign.get("kind", "scholarship")
    else:
        raise HTTPException(400, "scholarship_id or carnival_id required")

    clean_email = payload.email.lower().strip()
    clean_phone = payload.phone.strip()

    # Duplicate check within the same campaign or carnival
    dup_query: Dict[str, Any] = {"$or": [{"email": clean_email}, {"phone": clean_phone}]}
    if payload.carnival_id:
        dup_query["carnival_id"] = payload.carnival_id
    else:
        dup_query["scholarship_id"] = payload.scholarship_id

    existing_app = await db.scholarship_applications.find_one(dup_query)
    if existing_app:
        msg = f"An application already exists with this data (App No: {existing_app.get('application_no')})."
        raise HTTPException(status_code=400, detail=msg)

    selected_venue = _sanitize_venue(payload.venue)

    # If carnival: try to reserve the slot atomically BEFORE inserting the app
    if carnival:
        ok = await try_reserve_slot(db, carnival["id"], payload.chosen_date, payload.chosen_slot_time)
        if not ok:
            raise HTTPException(409, "This slot is now full or closed — please pick another")

    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["email"] = clean_email
    doc["phone"] = clean_phone
    doc["campaign_kind"] = campaign_kind
    doc["source"] = "self"

    for _ in range(10):
        candidate = str(random.randint(10000000, 99999999))
        if not await db.scholarship_applications.find_one({"application_no": candidate}):
            doc["application_no"] = candidate
            break
    else:
        doc["application_no"] = str(int(datetime.now(timezone.utc).timestamp() * 1000))[-8:]

    doc["status"] = "pending"
    if campaign:
        doc["scholarship_title"] = campaign.get("title", "")
        exam_date_str = campaign.get("exam_date", "TBA")
        exam_time_str = campaign.get("exam_time", "10:00 AM")
        title_for_email = campaign.get("title") or payload.target_exam
    else:
        doc["scholarship_title"] = carnival.get("title", "WATH Carnival")
        exam_date_str = payload.chosen_date
        exam_time_str = payload.chosen_slot_time
        title_for_email = carnival.get("title") or "WATH Carnival"
    doc["venue"] = selected_venue
    doc["created_at"] = now_iso()

    try:
        await db.scholarship_applications.insert_one(doc)
    except Exception:
        # Roll back the slot reservation if the insert fails
        if carnival:
            await release_slot(db, carnival["id"], payload.chosen_date, payload.chosen_slot_time)
        raise
    doc.pop("_id", None)
    doc["whatsapp_community_url"] = (campaign or carnival or {}).get("whatsapp_community_url")

    admit_pdf_bytes = None
    try:
        admit_pdf_bytes = admit_card_pdf(
            application_no=doc["application_no"],
            name=payload.name, phone=payload.phone, school=payload.school,
            standard=payload.standard, target_exam=payload.target_exam,
            exam_date=exam_date_str, venue=selected_venue,
            exam_time=exam_time_str,
            scholarship_title=title_for_email,
            father_name=payload.father_name, gender=payload.gender,
            dob=payload.dob, email=payload.email,
            address=payload.address, district=payload.district,
        )
    except Exception as e:
        logging.error(f"Failed to generate Admit Card PDF: {e}")

    background.add_task(
        email_scholarship_received, payload.email, payload.name, doc["application_no"], payload.target_exam, admit_pdf_bytes
    )
    background.add_task(
        email_admin_notification, f"New {campaign_kind} application: {payload.name}",
        f"<p><b>{payload.name}</b> from {payload.school} ({payload.standard}) applied for <b>{title_for_email}</b> at <b>{doc['venue']}</b>.<br/>App No: {doc['application_no']}<br/>Phone: {payload.phone}<br/>Slot: {exam_date_str} · {exam_time_str}</p>"
    )

    if admit_pdf_bytes:
        background.add_task(
            _safe_send_whatsapp_admit_card,
            phone=payload.phone,
            name=payload.name,
            application_no=doc["application_no"],
            scholarship_title=title_for_email,
            standard=payload.standard,
            exam_date=exam_date_str,
            exam_time=exam_time_str,
            venue=selected_venue,
            pdf_bytes=admit_pdf_bytes,
        )

    return doc

@api.put("/scholarship-applications/{application_no}")
async def update_scholarship_application(
    application_no: str,
    payload: ScholarshipApplicationUpdateIn,
    _admin = Depends(require_admin)
):
    """Update applicant details, including school venues and contact details."""
    existing_app = await db.scholarship_applications.find_one({"application_no": application_no.strip()})
    if not existing_app:
        existing_app = await db.scholarship_applications.find_one({"id": application_no.strip()})
        if not existing_app:
            raise HTTPException(status_code=404, detail="Scholarship application not found")

    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields provided to update")

    if "email" in update_data:
        update_data["email"] = update_data["email"].lower().strip()
    if "phone" in update_data:
        update_data["phone"] = update_data["phone"].strip()
    if "venue" in update_data:
        update_data["venue"] = _sanitize_venue(update_data["venue"])

    update_data["updated_at"] = now_iso()

    await db.scholarship_applications.update_one(
        {"_id": existing_app["_id"]},
        {"$set": update_data}
    )

    updated_doc = await db.scholarship_applications.find_one({"_id": existing_app["_id"]}, {"_id": 0})
    return updated_doc

@api.get("/scholarship-applications")
async def list_scholarship_apps(_admin = Depends(require_admin)):
    # Limit removed so the frontend matrix counts represent the entire applicant pool
    return await db.scholarship_applications.find({}, {"_id": 0}).sort("created_at", -1).to_list(None)

@api.post("/admin/scholarships/{scholarship_id}/notify-applicants")
async def notify_scholarship_applicants(
    scholarship_id: str, background: BackgroundTasks, _admin = Depends(require_admin)
):
    campaign = await db.scholarships.find_one({"id": scholarship_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Scholarship campaign not found")

    total_applicants = await db.scholarship_applications.count_documents({"scholarship_id": scholarship_id})
    if total_applicants == 0:
        raise HTTPException(status_code=400, detail="No applicants found for this scholarship")

    job_id = new_id()
    await db.bulk_jobs.insert_one({
        "id": job_id,
        "scholarship_id": scholarship_id,
        "status": "initializing",
        "total_rows": total_applicants,
        "processed": 0,
        "success": 0,
        "errors": 0,
        "recent_logs": [f"🚀 Starting WhatsApp broadcast for {total_applicants} applicants..."],
        "created_at": now_iso()
    })

    background.add_task(broadcast_scholarship_details, scholarship_id, job_id)

    return {
        "status": "accepted",
        "message": f"Notification broadcast started for {total_applicants} applicants.",
        "scholarship_id": scholarship_id,
        "job_id": job_id
    }

# ---------- Per-campaign Registrations Dashboard ----------
@api.get("/scholarships/{sid}/stats")
async def scholarship_stats(sid: str, request: Request, token: str | None = None):
    camp = await db.scholarships.find_one({"$or": [{"id": sid}, {"slug": sid}]}, {"_id": 0})
    if not camp:
        raise HTTPException(404, "Campaign not found")
    real_id = camp["id"]

    # Check authorization or examiner token
    is_authorized = False
    try:
        user = await get_current_user(request)
        if user and user.get("role") in ("admin", "super_admin"):
            is_authorized = True
    except HTTPException:
        pass

    if not is_authorized and token and camp.get("examiner_token") == token:
        is_authorized = True

    if not is_authorized:
        raise HTTPException(401, "Authentication required to view stats")

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)
    prev_week_start = today_start - timedelta(days=14)

    apps = await db.scholarship_applications.find(
        {"scholarship_id": real_id}, {"_id": 0, "venue": 1, "created_at": 1}
    ).to_list(None)

    def _parse(ts):
        if not ts:
            return None
        try:
            return datetime.fromisoformat(ts.replace("Z", "+00:00")) if isinstance(ts, str) else ts
        except Exception:
            return None

    venues_seed = list(camp.get("available_venues") or [])
    venues = {v: {"venue": v, "total": 0, "today": 0, "last_7_days": 0} for v in venues_seed}

    total = 0
    this_week = 0
    prev_week = 0
    for a in apps:
        total += 1
        v = a.get("venue") or "—"
        if v not in venues:
            venues[v] = {"venue": v, "total": 0, "today": 0, "last_7_days": 0}
        venues[v]["total"] += 1
        ts = _parse(a.get("created_at"))
        if ts:
            if ts.tzinfo is None:
                ts = ts.replace(tzinfo=timezone.utc)
            if ts >= today_start:
                venues[v]["today"] += 1
            if ts >= week_start:
                venues[v]["last_7_days"] += 1
                this_week += 1
            elif ts >= prev_week_start:
                prev_week += 1

    by_venue = sorted(venues.values(), key=lambda x: (-x["total"], x["venue"]))
    top_venue = by_venue[0]["venue"] if by_venue and by_venue[0]["total"] > 0 else None

    if prev_week > 0:
        wow_growth_pct = round(((this_week - prev_week) / prev_week) * 100, 2)
    else:
        wow_growth_pct = 100.0 if this_week > 0 else 0.0

    return {
        "campaign": {"id": camp["id"], "slug": camp.get("slug"), "title": camp.get("title"), "exam_date": camp.get("exam_date")},
        "total_registrations": total,
        "wow_growth_pct": wow_growth_pct,
        "this_week": this_week,
        "prev_week": prev_week,
        "by_venue": by_venue,
        "top_venue": top_venue,
        "as_of": now.isoformat(),
    }


@api.get("/scholarship-applications/mine")
async def my_scholarship_applications(user: dict = Depends(get_current_user)):
    q = {"$or": [{"email": user.get("email")}]}
    if user.get("phone"):
        q["$or"].append({"phone": user["phone"]})
    apps = await db.scholarship_applications.find(q, {"_id": 0}).sort("created_at", -1).to_list(None)
    for a in apps:
        if not a.get("result_published"):
            for k in ("result_marks_obtained", "result_total_marks", "result_rank", "result_percentile", "result_scholarship_percentage", "result_remarks"):
                a.pop(k, None)
    return apps


@api.get("/scholarship-applications/{application_no}")
async def get_scholarship_app_by_no(application_no: str, phone: Optional[str] = Query(None)):
    if not phone:
        raise HTTPException(400, "Phone number is required to view this application")
    app_doc = await db.scholarship_applications.find_one(
        {"application_no": application_no.strip(), "phone": phone.strip()},
        {"_id": 0}
    )
    if not app_doc:
        raise HTTPException(404, "No application found with this number and phone")
    if not app_doc.get("result_published"):
        for k in ("result_marks_obtained", "result_total_marks", "result_rank", "result_percentile", "result_scholarship_percentage", "result_remarks"):
            app_doc.pop(k, None)
    return app_doc

@api.put("/scholarship-applications/{aid}/status")
async def update_scholarship_status(aid: str, status: str = Query(...), _admin = Depends(require_admin)):
    if status not in ("pending", "approved", "rejected"):
        raise HTTPException(400, "Invalid status")
    await db.scholarship_applications.update_one({"id": aid}, {"$set": {"status": status}})
    return {"ok": True}

# ---------- Scholarship Result management ----------
@api.put("/scholarship-applications/{aid}/result")
async def set_scholarship_result(aid: str, payload: ScholarshipResultIn, background: BackgroundTasks, _admin = Depends(require_admin)):
    app_doc = await db.scholarship_applications.find_one({"id": aid}, {"_id": 0})
    if not app_doc:
        raise HTTPException(404, "Application not found")
    pct = max(0, min(100, payload.scholarship_percentage))
    update = {
        "result_marks_obtained": payload.marks_obtained,
        "result_total_marks": payload.total_marks,
        "result_rank": payload.rank,
        "result_percentile": payload.percentile,
        "result_scholarship_percentage": pct,
        "result_remarks": payload.remarks,
        "result_published": bool(payload.publish),
    }
    was_published = bool(app_doc.get("result_published"))
    if payload.publish:
        update["result_published_at"] = now_iso()
    await db.scholarship_applications.update_one({"id": aid}, {"$set": update})
    if payload.publish and not was_published and app_doc.get("email"):
        front = os.environ.get("FRONTEND_URL", "").rstrip("/")
        result_url = (f"{front}/api/scholarship-applications/{app_doc['application_no']}/result-card"
                      f"?phone={app_doc.get('phone','')}") if front else None
        background.add_task(
            email_scholarship_result_published,
            app_doc["email"], app_doc.get("name", ""), app_doc["application_no"],
            pct, payload.marks_obtained, payload.total_marks, payload.rank,
            result_url,
        )
    return {"ok": True, **update}

@api.post("/scholarship-applications/lookup")
async def lookup_scholarship(payload: ScholarshipLookupIn):
    app_doc = await db.scholarship_applications.find_one(
        {"application_no": payload.application_no.strip(), "phone": payload.phone.strip()},
        {"_id": 0}
    )
    if not app_doc:
        raise HTTPException(404, "No application found with this number and phone")
    if not app_doc.get("result_published"):
        for k in ("result_marks_obtained", "result_total_marks", "result_rank", "result_percentile", "result_scholarship_percentage", "result_remarks"):
            app_doc.pop(k, None)
    return app_doc

# ---------- School helpers ----------

def _parse_school_excel(file_bytes: bytes) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    headers = [str((c.value or "").strip()).lower() for c in ws[1]]
    required = {"name", "mobile", "current class", "course"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(sorted(missing))}")
    name_idx = headers.index("name")
    mobile_idx = headers.index("mobile")
    class_idx = headers.index("current class")
    course_idx = headers.index("course")
    ALLOWED_CLASSES = {"7th class", "8th class", "9th class", "10th class", "11th class", "12th class"}
    ALLOWED_COURSES = {"foundation", "neet", "iit jee"}
    rows = []
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        name = str(row[name_idx] or "").strip()
        mobile = str(row[mobile_idx] or "").strip()
        cur_class = str(row[class_idx] or "").strip().lower()
        course = str(row[course_idx] or "").strip().lower()
        if not name or not mobile or not cur_class or not course:
            errors.append(f"Row {i}: missing required values")
            continue
        if cur_class not in ALLOWED_CLASSES:
            errors.append(f"Row {i}: invalid class '{row[class_idx]}'")
            continue
        if course not in ALLOWED_COURSES:
            errors.append(f"Row {i}: invalid course '{row[course_idx]}'")
            continue
        rows.append({
            "name": name,
            "mobile": mobile,
            "current_class": cur_class.title(),
            "course": course.title() if course != "iit jee" else "IIT JEE",
        })
    return rows, errors

# ---------- School routes ----------

@api.post("/school/register")
async def school_register(payload: RegisterIn, response: Response):
    if not payload.school_name:
        raise HTTPException(400, "school_name is required for school registration")
    return await register(payload, response)

async def _get_school_visit_or_404(visit_id: str):
    visit = await db.school_visits.find_one({"id": visit_id}, {"_id": 0})
    if not visit:
        raise HTTPException(404, "Visit request not found")
    return visit

@api.post("/school/upload-students")
async def school_upload_students(
    scholarship_id: str = Form(...),
    file: UploadFile = File(...),
    school: dict = Depends(require_school),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are allowed")
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(400, "File size must be under 5MB")
    try:
        rows, parse_errors = _parse_school_excel(contents)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to parse Excel: {e}")
    if len(rows) > 500:
        raise HTTPException(400, "Maximum 500 student rows allowed per upload")
    campaign = await db.scholarships.find_one({"id": scholarship_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(404, "Scholarship campaign not found")
    if not campaign.get("active"):
        raise HTTPException(400, "Scholarship campaign is closed")
    created = 0
    skipped = 0
    for row in rows:
        dup_query = {"$or": [{"email": row["mobile"] + "@school.local"}, {"phone": row["mobile"]}]}
        dup_query["scholarship_id"] = scholarship_id
        existing = await db.scholarship_applications.find_one(dup_query)
        if existing:
            skipped += 1
            continue
        student_doc = {
            "id": new_id(),
            "name": row["name"],
            "email": f"{row['mobile']}@school.local",
            "phone": row["mobile"],
            "school": school.get("school_name") or school.get("name") or "",
            "standard": row["current_class"],
            "target_exam": row["course"],
            "city": school.get("district") or "",
            "scholarship_id": scholarship_id,
            "venue": campaign.get("available_venues", ["90 FT"])[0] if campaign.get("available_venues") else "90 FT",
            "address": school.get("address"),
            "district": school.get("district"),
            "source": "school",
            "school_id": school.get("id"),
            "school_name": school.get("school_name") or school.get("name") or "",
            "campaign_kind": "scholarship",
            "status": "pending",
            "scholarship_title": campaign.get("title", ""),
            "created_at": now_iso(),
        }
        for _ in range(10):
            candidate = str(random.randint(10000000, 99999999))
            if not await db.scholarship_applications.find_one({"application_no": candidate}):
                student_doc["application_no"] = candidate
                break
        else:
            student_doc["application_no"] = str(int(datetime.now(timezone.utc).timestamp() * 1000))[-8:]
        await db.scholarship_applications.insert_one(student_doc)
        created += 1
    return SchoolBulkRegisterResult(
        processed=len(rows),
        created=created,
        skipped=skipped,
        errors=parse_errors,
    ).model_dump()

@api.post("/school/visit-request")
async def school_visit_request(payload: SchoolVisitIn, school: dict = Depends(require_school)):
    try:
        pref_date = datetime.strptime(payload.preferred_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD")
    today = datetime.now(timezone.utc).date()
    if pref_date <= today:
        raise HTTPException(400, "Preferred date must be in the future")
    campaign = await db.scholarships.find_one({"id": payload.scholarship_id}, {"_id": 0})
    if not campaign:
        raise HTTPException(404, "Scholarship campaign not found")
    existing = await db.school_visits.find_one({"school_id": school["id"], "scholarship_id": payload.scholarship_id})
    if existing:
        raise HTTPException(400, "You have already submitted a visit request for this campaign")
    count = await db.school_visits.count_documents({
        "preferred_date": payload.preferred_date,
        "status": {"$in": ["pending", "approved"]},
    })
    if count >= 2:
        raise HTTPException(400, "This date already has 2 schools scheduled. Please choose another date")
    visit = {
        "id": new_id(),
        "school_id": school["id"],
        "school_name": school.get("school_name") or school.get("name") or "",
        "scholarship_id": payload.scholarship_id,
        "preferred_date": payload.preferred_date,
        "preferred_slot_time": payload.preferred_slot_time,
        "status": "pending",
        "admin_notes": payload.notes or "",
        "created_at": now_iso(),
    }
    await db.school_visits.insert_one(visit)
    visit.pop("_id", None)
    return visit

@api.get("/school/my-visits")
async def school_my_visits(school: dict = Depends(require_school)):
    visits = await db.school_visits.find({"school_id": school["id"]}, {"_id": 0}).sort("created_at", -1).to_list(None)
    for v in visits:
        v.setdefault("admin_notes", "")
    return visits

@api.get("/admin/school-visits")
async def admin_list_school_visits(
    status: Optional[str] = Query(None),
    date: Optional[str] = Query(None),
    _admin = Depends(require_admin),
):
    q: Dict[str, Any] = {}
    if status:
        q["status"] = status
    if date:
        q["preferred_date"] = date
    visits = await db.school_visits.find(q, {"_id": 0}).sort("created_at", -1).to_list(None)
    for v in visits:
        v.setdefault("admin_notes", "")
    return visits

@api.put("/admin/school-visits/{visit_id}")
async def admin_update_school_visit(
    visit_id: str,
    payload: Dict[str, Any],
    _admin = Depends(require_admin),
):
    visit = await _get_school_visit_or_404(visit_id)
    allowed_statuses = {"pending", "approved", "rejected"}
    update_fields = {}
    if "status" in payload:
        if payload["status"] not in allowed_statuses:
            raise HTTPException(400, "Invalid status")
        update_fields["status"] = payload["status"]
    if "admin_notes" in payload:
        update_fields["admin_notes"] = payload.get("admin_notes") or ""
    if not update_fields:
        raise HTTPException(400, "No updatable fields provided")
    await db.school_visits.update_one({"id": visit_id}, {"$set": update_fields})
    updated = await _get_school_visit_or_404(visit_id)
    return updated

@api.get("/admin/school-visits/availability")
async def admin_school_visit_availability(date: str = Query(...), _admin = Depends(require_admin)):
    try:
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD")
    count = await db.school_visits.count_documents({
        "preferred_date": date,
        "status": {"$in": ["pending", "approved"]},
    })
    return {"date": date, "current_count": count, "max": 2, "available": count < 2}

@api.get("/school/upload-template")
async def school_upload_template(school: dict = Depends(require_school)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Name", "Mobile", "Current Class", "Course"])
    ws.append(["Rahul Kumar", "9876543210", "10th Class", "NEET"])
    ws.append(["Ayesha Singh", "9876543211", "12th Class", "IIT JEE"])
    for col in ws.columns:
        for cell in col:
            if cell.value:
                cell.font = openpyxl.styles.Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="school_students_template.xlsx"'}
    )

@api.get("/admin/school-applications")
async def admin_list_school_applications(
    scholarship_id: Optional[str] = Query(None),
    school_id: Optional[str] = Query(None),
    _admin = Depends(require_admin),
):
    q: Dict[str, Any] = {"source": "school"}
    if scholarship_id:
        q["scholarship_id"] = scholarship_id
    if school_id:
        q["school_id"] = school_id
    rows = await db.scholarship_applications.find(q, {"_id": 0}).sort("created_at", -1).to_list(None)
    return rows

# ---------- Attendance (token-based, no login) ----------
async def _campaign_by_token(token: str):
    if not token:
        raise HTTPException(401, "Missing token")
    camp = await db.scholarships.find_one({"examiner_token": token}, {"_id": 0})
    if not camp:
        raise HTTPException(401, "Invalid examiner token")
    return camp

@api.get("/attendance/campaign")
async def attendance_campaign(token: str = Query(...)):
    camp = await _campaign_by_token(token)
    return {
        "id": camp["id"], "title": camp["title"],
        "exam_date": camp.get("exam_date"), "exam_time": camp.get("exam_time"),
        "available_venues": camp.get("available_venues") or [],
        "total_marks": camp.get("total_marks"),
    }

@api.get("/attendance/applications")
async def attendance_applications(token: str = Query(...), venue: Optional[str] = None):
    camp = await _campaign_by_token(token)
    q = {"scholarship_id": camp["id"]}
    if venue: q["venue"] = _sanitize_venue(venue)
    apps = await db.scholarship_applications.find(q, {"_id": 0}).sort("name", 1).to_list(None)
    appno_set = [a["application_no"] for a in apps]
    att_rows = await db.attendance.find({"scholarship_id": camp["id"], "application_no": {"$in": appno_set}}, {"_id": 0}).to_list(None)
    att_by = {a["application_no"]: a for a in att_rows}
    out = []
    for a in apps:
        rec = att_by.get(a["application_no"])
        out.append({
            "application_no": a["application_no"], "name": a["name"], "phone": a["phone"],
            "school": a.get("school"), "standard": a.get("standard"),
            "venue": _sanitize_venue(a.get("venue")),
            "attendance_status": (rec or {}).get("status"),
            "marked_at": (rec or {}).get("marked_at"),
        })
    return {"campaign_id": camp["id"], "venue": venue, "items": out,
            "marked_count": sum(1 for o in out if o["attendance_status"] == "present")}

@api.post("/attendance/mark")
async def attendance_mark(payload: AttendanceMarkIn):
    camp = await _campaign_by_token(payload.token)
    app_no = payload.application_no.strip()
    app_doc = await db.scholarship_applications.find_one(
        {"application_no": app_no, "scholarship_id": camp["id"]}, {"_id": 0}
    )
    if not app_doc:
        raise HTTPException(404, "Application not found for this campaign")
    rec = {
        "application_no": app_no,
        "scholarship_id": camp["id"],
        "venue": _sanitize_venue(payload.venue),
        "status": payload.status,
        "marked_at": now_iso(),
    }
    await db.attendance.update_one(
        {"application_no": app_no, "scholarship_id": camp["id"]},
        {"$set": rec}, upsert=True,
    )
    return {"ok": True, "application_no": app_no, "name": app_doc.get("name"), "status": payload.status, "marked_at": rec["marked_at"]}

@api.get("/admin/attendance/{sid}/export")
async def attendance_export(sid: str, _admin = Depends(require_admin)):
    camp = await db.scholarships.find_one({"id": sid}, {"_id": 0})
    if not camp:
        raise HTTPException(404, "Campaign not found")
        
    total_count = await db.scholarship_applications.count_documents({"scholarship_id": sid})
    if total_count == 0:
        raise HTTPException(404, "No scholarship applications found for this campaign")

    projection = {"application_no": 1, "name": 1, "phone": 1, "school": 1, "standard": 1, "venue": 1, "_id": 0}
    apps = await db.scholarship_applications.find({"scholarship_id": sid}, projection).to_list(None)
    
    att_rows = await db.attendance.find({"scholarship_id": sid}, {"_id": 0}).to_list(None)
    att = {a["application_no"]: a for a in att_rows}

    def _generate_attendance_excel():
        rows = []
        for a in apps:
            rec = att.get(a.get("application_no")) or {}
            rows.append({
                "application_no": str(a.get("application_no", "")),
                "name": str(a.get("name", "")),
                "phone": str(a.get("phone", "")),
                "school": str(a.get("school", "")),
                "standard": str(a.get("standard", "")),
                "venue": _sanitize_venue(a.get("venue")),
                "status": str(rec.get("status") or "absent"),
                "marked_at": str(rec.get("marked_at") or ""),
            })
        return export_excel(rows, "Attendance", f"attendance-{sid}.xlsx")

    return await asyncio.to_thread(_generate_attendance_excel)

# ---------- Scholarship results: template + bulk upload ----------
RESULT_HEADERS = ["application_no", "name", "school", "standard",
                  "marks_obtained", "total_marks", "rank", "percentile",
                  "scholarship_percentage", "remarks", "publish"]

@api.get("/admin/scholarships/{sid}/results-template")
async def results_template(sid: str, _admin = Depends(require_admin)):
    camp = await db.scholarships.find_one({"id": sid}, {"_id": 0, "total_marks": 1, "title": 1})
    if not camp:
        raise HTTPException(404, "Campaign not found")
        
    total_count = await db.scholarship_applications.count_documents({"scholarship_id": sid})
    if total_count == 0:
        raise HTTPException(404, "No scholarship applications found for this campaign")

    projection = {"application_no": 1, "name": 1, "school": 1, "standard": 1, "_id": 0}
    apps = await db.scholarship_applications.find({"scholarship_id": sid}, projection).sort("name", 1).to_list(None)
    total_marks_default = camp.get("total_marks") or 100

    def _generate_excel():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(RESULT_HEADERS)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            
        for a in apps:
            ws.append([
                str(a.get("application_no", "")),
                str(a.get("name", "")),
                str(a.get("school", "")),
                str(a.get("standard", "")),
                "", total_marks_default, "", "", "", "", "no",
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = await asyncio.to_thread(_generate_excel)
    file_size = buf.getbuffer().nbytes

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="results-template-{sid}.xlsx"',
            "Content-Length": str(file_size)
        }
    )
    
@api.post("/admin/scholarships/{sid}/bulk-results")
async def bulk_results(sid: str, background: BackgroundTasks,
                       file: UploadFile = File(...), _admin = Depends(require_admin)):
    camp = await db.scholarships.find_one({"id": sid}, {"_id": 0})
    if not camp:
        raise HTTPException(404, "Campaign not found")
    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel: {e}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Sheet is empty")
    headers = [str(h or "").strip().lower() for h in rows[0]]
    h_idx = {h: i for i, h in enumerate(headers)}
    required = ["application_no", "marks_obtained", "scholarship_percentage"]
    missing = [r for r in required if r not in h_idx]
    if missing:
        raise HTTPException(400, f"Missing columns: {missing}. Expected: {RESULT_HEADERS}")

    front = os.environ.get("FRONTEND_URL", "").rstrip("/")
    processed, published_count, errors = 0, 0, []
    for row_no, row in enumerate(rows[1:], start=2):
        try:
            app_no = (str(row[h_idx["application_no"]]).strip() if row[h_idx["application_no"]] is not None else "")
            if not app_no:
                continue
            marks = float(row[h_idx["marks_obtained"]] or 0)
            total = float(row[h_idx.get("total_marks", -1)] or camp.get("total_marks") or 100) if "total_marks" in h_idx else float(camp.get("total_marks") or 100)
            sch_pct = max(0, min(100, int(float(row[h_idx["scholarship_percentage"]] or 0))))
            rank = row[h_idx.get("rank", -1)] if "rank" in h_idx else None
            rank_v = int(rank) if rank not in (None, "", 0) else None
            perc = row[h_idx.get("percentile", -1)] if "percentile" in h_idx else None
            perc_v = float(perc) if perc not in (None, "") else None
            remarks = row[h_idx.get("remarks", -1)] if "remarks" in h_idx else None
            remarks_v = str(remarks).strip() if remarks not in (None, "") else None
            pub_raw = row[h_idx.get("publish", -1)] if "publish" in h_idx else "yes"
            publish = str(pub_raw).strip().lower() in ("yes", "true", "1", "y", "publish", "published")

            app_doc = await db.scholarship_applications.find_one(
                {"application_no": app_no, "scholarship_id": sid}, {"_id": 0}
            )
            if not app_doc:
                errors.append({"row": row_no, "app": app_no, "error": "not in this campaign"})
                continue

            was_published = bool(app_doc.get("result_published"))
            update = {
                "result_marks_obtained": marks,
                "result_total_marks": total,
                "result_rank": rank_v,
                "result_percentile": perc_v,
                "result_scholarship_percentage": sch_pct,
                "result_remarks": remarks_v,
                "result_published": publish,
            }
            if publish:
                update["result_published_at"] = now_iso()
                published_count += 1
            await db.scholarship_applications.update_one({"id": app_doc["id"]}, {"$set": update})
            processed += 1

            if publish and not was_published and app_doc.get("email"):
                result_url = (f"{front}/api/scholarship-applications/{app_no}/result-card"
                              f"?phone={app_doc.get('phone','')}") if front else None
                background.add_task(
                    email_scholarship_result_published,
                    app_doc["email"], app_doc.get("name", ""), app_no,
                    sch_pct, marks, total, rank_v, result_url,
                )
        except Exception as e:
            errors.append({"row": row_no, "app": app_no if 'app_no' in locals() else "", "error": str(e)})
    return {"processed": processed, "published": published_count, "errors": errors}

# ---------- Bulk Scholarship Registration Endpoint ----------

BULK_SCHOLARSHIP_HEADERS = [
    "full_name", "email", "phone", "class", "school_institute", "venue"
]

@api.get("/admin/scholarships/{sid}/bulk-register-template")
async def bulk_register_template(sid: str, _admin = Depends(require_admin)):
    camp = await db.scholarships.find_one({"id": sid}, {"_id": 0})
    if not camp:
        raise HTTPException(404, "Campaign not found")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bulk Registrations"
    ws.append(BULK_SCHOLARSHIP_HEADERS)
    sample_venue = (camp.get("available_venues") or ["90 FT"])[0]
    ws.append([
        "Aarav Sharma", "aarav@example.com", "9999900000",
        "Class 10", "DPS Srinagar", sample_venue,
    ])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"bulk-register-template-{sid}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- BULK UPLOAD ASYNC WORKER ----------

async def _process_bulk_file_bg(job_id: str, sid: str, file_path: str):
    try:
        campaign = await db.scholarships.find_one({"id": sid}, {"_id": 0})
        if not campaign:
            raise Exception("Campaign not found during processing.")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows or len(rows) < 2:
            raise Exception("Sheet is empty or missing data rows.")

        raw_headers = [str(h or "").strip().lower().replace(" ", "_").replace("/", "_") for h in rows[0]]
        header_map = {}
        for idx, h in enumerate(raw_headers):
            if "name" in h and "school" not in h and "institute" not in h: header_map["name"] = idx
            elif "email" in h: header_map["email"] = idx
            elif "phone" in h or "mobile" in h or "contact" in h: header_map["phone"] = idx
            elif "class" in h or "standard" in h or "grade" in h: header_map["standard"] = idx
            elif "school" in h or "institute" in h: header_map["school"] = idx
            elif "venue" in h or "location" in h or "center" in h: header_map["venue"] = idx

        for req in ("name", "email", "phone"):
            if req not in header_map:
                raise Exception(f"Missing required column: {req}")

        valid_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        total_rows = len(valid_rows)

        await db.bulk_jobs.update_one(
            {"id": job_id}, 
            {"$set": {"status": "processing", "total_rows": total_rows}}
        )

        for row_no, row in enumerate(valid_rows, start=2):
            try:
                def _cell(key: str, default: str = "") -> str:
                    idx = header_map.get(key)
                    if idx is None or idx >= len(row) or row[idx] is None: return default
                    return str(row[idx]).strip()

                name = _cell("name")
                email = _cell("email").lower()
                phone_raw = _cell("phone")
                phone = str(phone_raw).split(".")[0].strip() if phone_raw else ""
                standard = _cell("standard", "General")
                school = _cell("school", "N/A")
                venue = _sanitize_venue(_cell("venue"))

                if not name or not email or not phone:
                    raise ValueError(f"Missing required data for Name: {name}")

                existing = await db.scholarship_applications.find_one({
                    "scholarship_id": sid,
                    "$or": [{"email": email}, {"phone": phone}],
                })
                
                if existing:
                    log_msg = f"⚠️ Skipped {name} (Row {row_no}): Already registered."
                    await db.bulk_jobs.update_one({"id": job_id}, {
                        "$inc": {"processed": 1, "errors": 1},
                        "$push": {"recent_logs": {"$each": [log_msg], "$slice": -15}}
                    })
                    continue

                for _ in range(10):
                    candidate = str(random.randint(10000000, 99999999))
                    if not await db.scholarship_applications.find_one({"application_no": candidate}):
                        app_no = candidate
                        break
                else:
                    app_no = str(int(datetime.now(timezone.utc).timestamp() * 1000))[-8:]
                
                doc = {
                    "id": new_id(),
                    "application_no": app_no,
                    "scholarship_id": sid,
                    "scholarship_title": campaign.get("title", "Scholarship Test"),
                    "name": name, "email": email, "phone": phone,
                    "standard": standard, "school": school,
                    "target_exam": standard, "city": venue, "venue": venue,
                    "status": "approved",
                    "created_at": now_iso(),
                }
                await db.scholarship_applications.insert_one(doc)

                admit_pdf_bytes = None
                try:
                    admit_pdf_bytes = admit_card_pdf(
                        application_no=app_no, name=name, phone=phone, school=school,
                        standard=standard, target_exam=standard,
                        exam_date=campaign.get("exam_date", "TBA"), venue=venue,
                        exam_time=campaign.get("exam_time", "10:00 AM"),
                        scholarship_title=campaign.get("title", "Scholarship Test"),
                        father_name=doc.get("father_name"), gender=doc.get("gender"),
                        dob=doc.get("dob"), email=doc.get("email"),
                        address=doc.get("address"), district=doc.get("district"),
                    )
                except Exception as e:
                    logging.error(f"PDF gen failed for {app_no}: {e}")

                asyncio.create_task(
                    _run_maybe_async(
                        email_scholarship_received,
                        email, name, app_no, standard, admit_pdf_bytes
                    )
                )

                if admit_pdf_bytes:
                    asyncio.create_task(
                        _run_maybe_async(
                            send_whatsapp_admit_card,
                            phone=phone,
                            name=name,
                            application_no=app_no,
                            scholarship_title=campaign.get("title") or standard,
                            standard=standard,
                            exam_date=campaign.get("exam_date") or "TBA",
                            exam_time=campaign.get("exam_time", "10:00 AM"),
                            venue=venue,
                            pdf_bytes=admit_pdf_bytes,
                        )
                    )
                        
                log_msg = f"✅ Success: {name} - Registered & WhatsApp Queued"
                await db.bulk_jobs.update_one({"id": job_id}, {
                    "$inc": {"processed": 1, "success": 1},
                    "$push": {"recent_logs": {"$each": [log_msg], "$slice": -15}}
                })
                
                await asyncio.sleep(0.01)

            except Exception as row_err:
                log_msg = f"❌ Error row {row_no} ({name if 'name' in locals() else 'Unknown'}): {str(row_err)}"
                await db.bulk_jobs.update_one({"id": job_id}, {
                    "$inc": {"processed": 1, "errors": 1},
                    "$push": {"recent_logs": {"$each": [log_msg], "$slice": -15}}
                })

        await db.bulk_jobs.update_one({"id": job_id}, {"$set": {"status": "completed"}})

    except Exception as e:
        await db.bulk_jobs.update_one({"id": job_id}, {
            "$set": {"status": "failed"}, 
            "$push": {"recent_logs": {"$each": [f"💥 FATAL ERROR: {str(e)}"], "$slice": -15}}
        })
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@api.post("/admin/scholarships/{sid}/bulk-register")
async def bulk_register_scholarship(
    sid: str,
    background: BackgroundTasks,
    file: UploadFile = File(...),
    _admin = Depends(require_admin),
):
    campaign = await db.scholarships.find_one({"id": sid}, {"_id": 0})
    if not campaign:
        raise HTTPException(404, "Campaign not found")

    data = await file.read()
    if not data:
        raise HTTPException(400, "Empty file")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {e}")

    if not rows or len(rows) < 2:
        return {"registered": 0, "skipped": 0, "total_rows": 0, "errors": []}

    raw_headers = [str(h or "").strip().lower().replace(" ", "_").replace("/", "_") for h in rows[0]]
    header_map = {}
    for idx, h in enumerate(raw_headers):
        if "name" in h and "school" not in h and "institute" not in h: header_map["name"] = idx
        elif "email" in h: header_map["email"] = idx
        elif "phone" in h or "mobile" in h or "contact" in h: header_map["phone"] = idx
        elif "class" in h or "standard" in h or "grade" in h: header_map["standard"] = idx
        elif "school" in h or "institute" in h: header_map["school"] = idx
        elif "venue" in h or "location" in h or "center" in h: header_map["venue"] = idx

    registered = 0
    skipped = 0
    errors = []

    valid_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]

    for row_no, row in enumerate(valid_rows, start=2):
        def _cell(key: str, default: str = "") -> str:
            idx = header_map.get(key)
            if idx is None or idx >= len(row) or row[idx] is None: return default
            return str(row[idx]).strip()

        name = _cell("name")
        email = _cell("email").lower()
        phone_raw = _cell("phone")
        phone = str(phone_raw).split(".")[0].strip() if phone_raw else ""
        standard = _cell("standard", "General")
        school = _cell("school", "N/A")
        venue = _sanitize_venue(_cell("venue"))

        if not name or not email or not phone:
            skipped += 1
            errors.append({
                "row": row_no,
                "reason": "missing_required",
                "message": f"Row {row_no} is missing required fields (name, email, or phone)."
            })
            continue

        existing = await db.scholarship_applications.find_one({
            "scholarship_id": sid,
            "$or": [{"email": email}, {"phone": phone}],
        })

        if existing:
            skipped += 1
            errors.append({
                "row": row_no,
                "reason": "duplicate",
                "application_no": existing.get("application_no"),
                "message": f"Row {row_no} ({name}): Email or phone already registered."
            })
            continue

        for _ in range(10):
            candidate = str(random.randint(10000000, 99999999))
            if not await db.scholarship_applications.find_one({"application_no": candidate}):
                app_no = candidate
                break
        else:
            app_no = str(int(datetime.now(timezone.utc).timestamp() * 1000))[-8:]

        doc = {
            "id": new_id(),
            "application_no": app_no,
            "scholarship_id": sid,
            "scholarship_title": campaign.get("title", "Scholarship Test"),
            "name": name, "email": email, "phone": phone,
            "standard": standard, "school": school,
            "target_exam": standard, "city": venue, "venue": venue,
            "status": "approved",
            "created_at": now_iso(),
        }
        await db.scholarship_applications.insert_one(doc)
        registered += 1

        admit_pdf_bytes = None
        try:
            admit_pdf_bytes = admit_card_pdf(
                application_no=app_no, name=name, phone=phone, school=school,
                standard=standard, target_exam=standard,
                exam_date=campaign.get("exam_date", "TBA"), venue=venue,
                exam_time=campaign.get("exam_time", "10:00 AM"),
                scholarship_title=campaign.get("title", "Scholarship Test"),
                father_name=doc.get("father_name"), gender=doc.get("gender"),
                dob=doc.get("dob"), email=doc.get("email"),
                address=doc.get("address"), district=doc.get("district"),
            )
        except Exception as e:
            logging.error(f"PDF gen failed for {app_no}: {e}")

        background.add_task(
            email_scholarship_received, email, name, app_no, standard, admit_pdf_bytes
        )

        if admit_pdf_bytes:
            background.add_task(
                _safe_send_whatsapp_admit_card,
                phone=phone,
                name=name,
                application_no=app_no,
                scholarship_title=campaign.get("title") or standard,
                standard=standard,
                exam_date=campaign.get("exam_date") or "TBA",
                exam_time=campaign.get("exam_time", "10:00 AM"),
                venue=venue,
                pdf_bytes=admit_pdf_bytes,
            )

    job_id = new_id()
    job_doc = {
        "id": job_id,
        "scholarship_id": sid,
        "status": "completed",
        "total_rows": len(valid_rows),
        "processed": len(valid_rows),
        "success": registered,
        "errors": skipped,
        "recent_logs": [f"✅ Processed {len(valid_rows)} rows: {registered} registered, {skipped} skipped."],
        "created_at": now_iso()
    }
    await db.bulk_jobs.insert_one(job_doc)

    return {
        "status": "accepted",
        "job_id": job_id,
        "registered": registered,
        "skipped": skipped,
        "total_rows": len(valid_rows),
        "errors": errors
    }


@api.get("/admin/bulk-jobs/{job_id}")
async def get_bulk_job_status(job_id: str, _admin = Depends(require_admin)):
    job = await db.bulk_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    return job

# ---------- Enrollments ----------
@api.post("/enrollments")
async def create_enrollment(payload: EnrollmentIn, request: Request, background: BackgroundTasks):
    course = await db.courses.find_one({"id": payload.course_id}, {"_id": 0})
    if not course:
        raise HTTPException(404, "Course not found")
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["status"] = "pending"
    doc["receipt_no"] = "UAC-ENR-" + str(uuid.uuid4().int)[:8]
    doc["created_at"] = now_iso()
    try:
        user = await get_current_user(request)
        doc["user_id"] = user["id"]
    except Exception:
        doc["user_id"] = None
    await db.enrollments.insert_one(doc)
    doc.pop("_id", None)
    background.add_task(email_enrollment_received, payload.email, payload.name, doc["receipt_no"], course["title"], payload.center)
    background.add_task(email_admin_notification, f"New enrollment: {payload.name}",
                       f"<p><b>{payload.name}</b> ({payload.email}, {payload.phone}) enrolled for <b>{course['title']}</b> at <b>{payload.center}</b>.<br/>Receipt: {doc['receipt_no']}</p>")
    return doc

@api.get("/enrollments")
async def list_enrollments(_admin = Depends(require_admin)):
    return await db.enrollments.find({}, {"_id": 0}).sort("created_at", -1).to_list(None)

@api.get("/enrollments/mine")
async def my_enrollments(user: dict = Depends(get_current_user)):
    return await db.enrollments.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(None)

@api.put("/enrollments/{eid}/status")
async def update_enrollment_status(eid: str, status: str = Query(...), _admin = Depends(require_admin)):
    if status not in ("pending", "approved", "rejected"):
        raise HTTPException(400, "Invalid status")
    await db.enrollments.update_one({"id": eid}, {"$set": {"status": status}})
    return {"ok": True}

# ---------- Jobs ----------
@api.get("/jobs")
async def list_jobs():
    return await db.jobs.find({"active": True}, {"_id": 0}).sort("created_at", -1).to_list(None)

@api.get("/jobs/all")
async def list_all_jobs(_admin = Depends(require_admin)):
    return await db.jobs.find({}, {"_id": 0}).sort("created_at", -1).to_list(None)

@api.post("/jobs")
async def create_job(payload: JobIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    await db.jobs.insert_one(doc)
    if doc.get("is_featured"):
        await _clear_featured_except("jobs", doc["id"])
    doc.pop("_id", None)
    return doc

@api.put("/jobs/{jid}")
async def update_job(jid: str, payload: JobIn, _admin = Depends(require_admin)):
    data = payload.model_dump()
    await db.jobs.update_one({"id": jid}, {"$set": data})
    if data.get("is_featured"):
        await _clear_featured_except("jobs", jid)
    return await db.jobs.find_one({"id": jid}, {"_id": 0})

@api.delete("/jobs/{jid}")
async def delete_job(jid: str, _admin = Depends(require_admin)):
    await db.jobs.delete_one({"id": jid})
    return {"ok": True}

@api.post("/job-applications")
async def apply_job(payload: JobApplicationIn, background: BackgroundTasks):
    job = await db.jobs.find_one({"id": payload.job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["status"] = "received"
    doc["created_at"] = now_iso()
    await db.job_applications.insert_one(doc)
    doc.pop("_id", None)
    background.add_task(email_job_app_received, payload.email, payload.name, job["title"])
    background.add_task(email_admin_notification, f"New job application: {payload.name}",
                       f"<p><b>{payload.name}</b> ({payload.email}, {payload.phone}) applied for <b>{job['title']}</b>.<br/>Qualification: {payload.qualification}<br/>Experience: {payload.experience}<br/>Resume: {payload.resume_url or '—'}</p>")
    return doc

@api.get("/job-applications")
async def list_job_apps(_admin = Depends(require_admin)):
    return await db.job_applications.find({}, {"_id": 0}).sort("created_at", -1).to_list(None)

@api.put("/job-applications/{aid}/status")
async def update_job_app_status(aid: str, status: str = Query(...), _admin = Depends(require_admin)):
    if status not in ("received", "shortlisted", "rejected", "hired"):
        raise HTTPException(400, "Invalid status")
    await db.job_applications.update_one({"id": aid}, {"$set": {"status": status}})
    return {"ok": True}

# ---------- Notices ----------
@api.get("/notices")
async def list_notices():
    return await db.notices.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)

@api.post("/notices")
async def create_notice(payload: NoticeIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    await db.notices.insert_one(doc)
    if doc.get("is_featured"):
        await _clear_featured_except("notices", doc["id"])
    doc.pop("_id", None)
    return doc

@api.put("/notices/{nid}")
async def update_notice(nid: str, payload: NoticeIn, _admin = Depends(require_admin)):
    data = payload.model_dump()
    await db.notices.update_one({"id": nid}, {"$set": data})
    if data.get("is_featured"):
        await _clear_featured_except("notices", nid)
    return await db.notices.find_one({"id": nid}, {"_id": 0})

@api.delete("/notices/{nid}")
async def delete_notice(nid: str, _admin = Depends(require_admin)):
    await db.notices.delete_one({"id": nid})
    return {"ok": True}

# ---------- Centers ----------
@api.get("/centers")
async def list_centers():
    return await db.centers.find({}, {"_id": 0}).to_list(100)

@api.post("/centers")
async def create_center(payload: CenterIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    await db.centers.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/centers/{cid}")
async def update_center(cid: str, payload: CenterIn, _admin = Depends(require_admin)):
    await db.centers.update_one({"id": cid}, {"$set": payload.model_dump()})
    return await db.centers.find_one({"id": cid}, {"_id": 0})

@api.delete("/centers/{cid}")
async def delete_center(cid: str, _admin = Depends(require_admin)):
    await db.centers.delete_one({"id": cid})
    return {"ok": True}

# ---------- Results ----------
@api.get("/results")
async def list_results():
    return await db.results.find({}, {"_id": 0}).sort("year", -1).to_list(200)

@api.post("/results")
async def create_result(payload: ResultIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    await db.results.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/results/{rid}")
async def update_result(rid: str, payload: ResultIn, _admin = Depends(require_admin)):
    await db.results.update_one({"id": rid}, {"$set": payload.model_dump()})
    return await db.results.find_one({"id": rid}, {"_id": 0})

@api.delete("/results/{rid}")
async def delete_result(rid: str, _admin = Depends(require_admin)):
    await db.results.delete_one({"id": rid})
    return {"ok": True}

# ---------- Testimonials ----------
@api.get("/testimonials")
async def list_testimonials():
    return await db.testimonials.find({}, {"_id": 0}).to_list(50)

@api.post("/testimonials")
async def create_testimonial(payload: TestimonialIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    await db.testimonials.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/testimonials/{tid}")
async def update_testimonial(tid: str, payload: TestimonialIn, _admin = Depends(require_admin)):
    await db.testimonials.update_one({"id": tid}, {"$set": payload.model_dump()})
    return await db.testimonials.find_one({"id": tid}, {"_id": 0})

@api.delete("/testimonials/{tid}")
async def delete_testimonial(tid: str, _admin = Depends(require_admin)):
    await db.testimonials.delete_one({"id": tid})
    return {"ok": True}

# ---------- Gallery ----------
@api.get("/gallery")
async def list_gallery():
    return await db.gallery.find({}, {"_id": 0}).sort("order", 1).to_list(200)

@api.get("/admin/gallery")
async def list_admin_gallery(_admin = Depends(require_admin)):
    return await db.gallery.find({}, {"_id": 0}).sort("order", 1).to_list(200)

@api.post("/admin/gallery")
async def create_gallery_item(payload: GalleryItemIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    await db.gallery.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/admin/gallery/{gid}")
async def update_gallery_item(gid: str, payload: GalleryItemIn, _admin = Depends(require_admin)):
    await db.gallery.update_one({"id": gid}, {"$set": payload.model_dump()})
    return await db.gallery.find_one({"id": gid}, {"_id": 0})

@api.delete("/admin/gallery/{gid}")
async def delete_gallery_item(gid: str, _admin = Depends(require_admin)):
    await db.gallery.delete_one({"id": gid})
    return {"ok": True}

# ---------- Blog ----------
class PostIn(BaseModel):
    title: str
    slug: str
    excerpt: Optional[str] = None
    content: str
    category: Optional[str] = None
    tags: List[str] = []
    author: str = "Admin"
    featured_image_url: Optional[str] = None
    image_alt: Optional[str] = None
    og_image_url: Optional[str] = None
    meta_title: Optional[str] = None
    meta_description: Optional[str] = None
    status: str = "draft"
    visibility: str = "public"
    published_at: Optional[str] = None

@api.get("/posts")
async def list_posts():
    return await db.posts.find({"status": "published", "visibility": "public"}, {"_id": 0}).sort("published_at", -1).to_list(100)

@api.get("/posts/{slug}")
async def get_post(slug: str):
    post = await db.posts.find_one({"slug": slug, "status": "published", "visibility": "public"}, {"_id": 0})
    if not post:
        raise HTTPException(404, "Post not found")
    return post

@api.get("/admin/posts")
async def list_admin_posts(_admin = Depends(require_admin)):
    return await db.posts.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)

@api.post("/admin/posts")
async def create_post(payload: PostIn, _admin = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    if not doc.get("published_at") and doc.get("status") == "published":
        doc["published_at"] = now_iso()
    await db.posts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/admin/posts/{pid}")
async def update_post(pid: str, payload: PostIn, _admin = Depends(require_admin)):
    await db.posts.update_one({"id": pid}, {"$set": payload.model_dump()})
    return await db.posts.find_one({"id": pid}, {"_id": 0})

@api.delete("/admin/posts/{pid}")
async def delete_post(pid: str, _admin = Depends(require_admin)):
    await db.posts.delete_one({"id": pid})
    return {"ok": True}

# ---------- Contact ----------
@api.post("/contact")
async def contact(payload: ContactIn, request: Request):
    client_ip = request.client.host if request.client else "unknown"
    check_rate_limit(f"contact:{client_ip}", max_requests=10, window_seconds=60)
    doc = payload.model_dump()
    doc["id"] = new_id()
    doc["created_at"] = now_iso()
    doc["status"] = "new"
    await db.inquiries.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.get("/inquiries")
async def list_inquiries(_admin = Depends(require_admin)):
    return await db.inquiries.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)

# ---------- Admin Dashboard summary ----------
@api.get("/admin/summary")
async def admin_summary(_admin = Depends(require_admin)):
    return {
        "total_students": await db.users.count_documents({"role": "student"}),
        "total_courses": await db.courses.count_documents({}),
        "total_enrollments": await db.enrollments.count_documents({}),
        "pending_enrollments": await db.enrollments.count_documents({"status": "pending"}),
        "total_scholarship_apps": await db.scholarship_applications.count_documents({}),
        "total_job_apps": await db.job_applications.count_documents({}),
        "total_inquiries": await db.inquiries.count_documents({}),
        "total_jobs": await db.jobs.count_documents({}),
    }

# ---------- File Upload & Download ----------
@api.post("/upload")
async def upload(request: Request, file: UploadFile = File(...)):
    client_ip = request.client.host if request.client else "unknown"
    check_rate_limit(f"upload:{client_ip}", max_requests=20, window_seconds=60)
    ctype = file.content_type or "application/octet-stream"
    if ctype not in ALLOWED_UPLOAD_TYPES:
        raise HTTPException(415, f"Unsupported type {ctype}. Allowed: pdf, jpg, png, webp, mp4, webm, mov.")
    ext = ALLOWED_UPLOAD_TYPES[ctype]
    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "File too large (max 50 MB)")
    if len(data) == 0:
        raise HTTPException(400, "Empty file")
    file_id = new_id()
    path = f"{APP_NAME}/uploads/{file_id}.{ext}"
    try:
        result = await put_object(path, data, ctype)
    except RuntimeError as e:
        msg = str(e)
        if "not initialised" in msg.lower():
            raise HTTPException(500, "File upload is currently unavailable because object storage is not configured on the server. Please contact the administrator.")
        raise HTTPException(500, f"Upload failed: {e}")
    except Exception as e:
        raise HTTPException(500, f"Upload failed: {e}")
    record = {
        "id": file_id,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": ctype,
        "size": result.get("size", len(data)),
        "is_deleted": False,
        "created_at": now_iso(),
    }
    await db.files.insert_one(record)
    record.pop("_id", None)
    record["url"] = f"/api/files/{file_id}"
    return record

@api.get("/files/{file_id}")
async def get_file(file_id: str):
    record = await db.files.find_one({"id": file_id, "is_deleted": False}, {"_id": 0})
    if not record:
        raise HTTPException(404, "File not found")
    try:
        data, ctype = await get_object(record["storage_path"])
        filename = record.get("original_filename", f"{file_id}.bin")
        return Response(
            content=data,
            media_type=ctype,
            headers={"Content-Disposition": f'inline; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(404, f"File retrieval failed: {e}")

# ---------- Scholarship admit card PDF ----------
@api.get("/scholarship-applications/{application_no}/admit-card")
async def admit_card(application_no: str, phone: Optional[str] = Query(None), request: Request = None):
    is_admin = False
    try:
        if request:
            user = await get_current_user(request)
            is_admin = bool(user and user.get("role") in ("admin", "super_admin"))
    except HTTPException:
        is_admin = False

    if is_admin:
        app_doc = await db.scholarship_applications.find_one({"application_no": application_no}, {"_id": 0})
    else:
        if not phone:
            raise HTTPException(400, "Phone number required to download admit card")
        app_doc = await db.scholarship_applications.find_one(
            {"application_no": application_no, "phone": phone.strip()}, {"_id": 0}
        )

    if not app_doc:
        raise HTTPException(404, "Application not found")

    # Resolve exam context: prefer carnival (if this app booked a slot), then scholarship campaign.
    carnival = None
    campaign = None
    if app_doc.get("carnival_id"):
        carnival = await db.wath_carnivals.find_one({"id": app_doc["carnival_id"]}, {"_id": 0})
    if app_doc.get("scholarship_id") and not carnival:
        campaign = await db.scholarships.find_one({"id": app_doc["scholarship_id"]}, {"_id": 0})
    if not carnival and not campaign:
        campaign = await db.scholarships.find_one({"active": True}, {"_id": 0}, sort=[("created_at", -1)])

    venue_name = _sanitize_venue(app_doc.get("venue") or (campaign or {}).get("venue") or app_doc.get("city"))
    if carnival:
        title = carnival.get("title") or "WATH Carnival"
        exam_date = app_doc.get("chosen_date") or (carnival.get("exam_dates") or [{}])[0].get("date", "TBA")
        exam_time = app_doc.get("chosen_slot_time") or "10:00 AM"
    else:
        title = (campaign or {}).get("title") or app_doc.get("scholarship_title", "Scholarship Test")
        exam_date = (campaign or {}).get("exam_date", "TBA")
        exam_time = (campaign or {}).get("exam_time", "10:00 AM")

    pdf_bytes = admit_card_pdf(
        application_no=application_no,
        name=app_doc.get("name", ""),
        phone=app_doc.get("phone", ""),
        school=app_doc.get("school", ""),
        standard=app_doc.get("standard", ""),
        target_exam=app_doc.get("target_exam", ""),
        exam_date=exam_date,
        venue=venue_name,
        exam_time=exam_time,
        scholarship_title=title,
        father_name=app_doc.get("father_name"),
        gender=app_doc.get("gender"),
        dob=app_doc.get("dob"),
        email=app_doc.get("email"),
        address=app_doc.get("address"),
        district=app_doc.get("district"),
    )
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="admit-card-{application_no}.pdf"'})

# ---------- Scholarship result card PDF ----------
@api.get("/scholarship-applications/{application_no}/result-card")
async def result_card(application_no: str, phone: Optional[str] = None):
    app_doc = await db.scholarship_applications.find_one({"application_no": application_no}, {"_id": 0})
    if not app_doc:
        raise HTTPException(404, "Application not found")
    if not app_doc.get("result_published"):
        raise HTTPException(403, "Result not yet published")
    if phone and app_doc.get("phone") != phone.strip():
        raise HTTPException(403, "Phone does not match application")
    campaign = None
    if app_doc.get("carnival_id"):
        car = await db.wath_carnivals.find_one({"id": app_doc["carnival_id"]}, {"_id": 0})
        if car:
            campaign = {"title": car.get("title") or "WATH Carnival"}
    if not campaign and app_doc.get("scholarship_id"):
        campaign = await db.scholarships.find_one({"id": app_doc["scholarship_id"]}, {"_id": 0})
    pdf_bytes = result_card_pdf(
        application_no=application_no,
        name=app_doc.get("name", ""),
        school=app_doc.get("school", ""),
        standard=app_doc.get("standard", ""),
        target_exam=app_doc.get("target_exam", ""),
        marks_obtained=app_doc.get("result_marks_obtained", 0),
        total_marks=app_doc.get("result_total_marks", 100),
        rank=app_doc.get("result_rank"),
        percentile=app_doc.get("result_percentile"),
        scholarship_percentage=app_doc.get("result_scholarship_percentage", 0),
        remarks=app_doc.get("result_remarks"),
        scholarship_title=(campaign or {}).get("title") or app_doc.get("scholarship_title", "Scholarship Test"),
    )
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="result-{application_no}.pdf"'})

@api.get("/admin/export/{kind}")
async def export(kind: str, _admin = Depends(require_admin)):
    mapping = {
        "enrollments": ("enrollments", "Enrollments"),
        "scholarship-applications": ("scholarship_applications", "Scholarships"),
        "job-applications": ("job_applications", "Jobs"),
        "inquiries": ("inquiries", "Inquiries"),
        "students": ("users", "Students"),
    }
    if kind not in mapping:
        raise HTTPException(404, "Unknown export kind")
    coll, name = mapping[kind]
    q = {"role": "student"} if kind == "students" else {}
    rows = await db[coll].find(q, {"_id": 0, "password_hash": 0}).to_list(None)
    return export_excel(rows, name, f"{kind}.xlsx")

# ---------- Seed ----------
async def seed():
    admin_email_raw = os.environ.get("ADMIN_EMAIL")
    admin_pwd = os.environ.get("ADMIN_PASSWORD")
    if not admin_email_raw or not admin_pwd:
        logging.warning("ADMIN_EMAIL / ADMIN_PASSWORD not set — skipping admin seed")
        return
    admin_email = admin_email_raw.lower().strip()
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": new_id(), "name": "Unacademy Admin", "email": admin_email,
            "phone": "+91-9999999999", "role": "admin",
            "password_hash": hash_password(admin_pwd),
            "created_at": now_iso(),
        })
    elif not verify_password(admin_pwd, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_pwd), "role": "admin"}})

    seed_marker = await db.system_meta.find_one({"key": "initial_seed"})
    if seed_marker is None:
        has_existing_data = (
            (await db.centers.count_documents({}) > 0)
            or (await db.courses.count_documents({}) > 0)
            or (await db.notices.count_documents({}) > 0)
            or (await db.jobs.count_documents({}) > 0)
            or (await db.testimonials.count_documents({}) > 0)
            or (await db.scholarships.count_documents({}) > 0)
        )
        if has_existing_data:
            await db.system_meta.insert_one({
                "key": "initial_seed", "completed_at": now_iso(), "method": "auto-detected",
            })
        else:
            await _run_initial_seed()
            await db.system_meta.insert_one({
                "key": "initial_seed", "completed_at": now_iso(), "method": "fresh-install",
            })

    async for sc in db.scholarships.find({"examiner_token": {"$exists": False}}, {"_id": 0, "id": 1}):
        await db.scholarships.update_one({"id": sc["id"]}, {"$set": {"examiner_token": uuid.uuid4().hex}})

    # ---------- Migration: backfill scholarships.kind ----------
    if not (await db.system_meta.find_one({"key": "kind_backfill_v1"})):
        # Anything titled WATH becomes kind=wath. All other rows become kind=scholarship.
        await db.scholarships.update_many(
            {"title": {"$regex": "WATH", "$options": "i"}, "kind": {"$exists": False}},
            {"$set": {"kind": "wath"}},
        )
        await db.scholarships.update_many(
            {"kind": {"$exists": False}},
            {"$set": {"kind": "scholarship"}},
        )
        # Backfill campaign_kind on existing applications
        wath_ids = [d["id"] async for d in db.scholarships.find({"kind": "wath"}, {"_id": 0, "id": 1})]
        if wath_ids:
            await db.scholarship_applications.update_many(
                {"scholarship_id": {"$in": wath_ids}, "campaign_kind": {"$exists": False}},
                {"$set": {"campaign_kind": "wath"}},
            )
        await db.scholarship_applications.update_many(
            {"campaign_kind": {"$exists": False}},
            {"$set": {"campaign_kind": "scholarship"}},
        )
        # Ensure default WATH page config exists (mode=exam)
        if not await db.system_meta.find_one({"key": "wath_page_config"}):
            await db.system_meta.insert_one({
                "key": "wath_page_config", "mode": "exam",
                "active_carnival_id": None, "disabled_message": None,
                "updated_at": now_iso(),
            })
        await db.system_meta.insert_one({"key": "kind_backfill_v1", "completed_at": now_iso()})
        logging.info("kind_backfill_v1 migration complete")

    await db.scholarship_applications.create_index(
        [("scholarship_id", 1), ("email", 1)]
    )
    await db.scholarship_applications.create_index(
        [("scholarship_id", 1), ("phone", 1)]
    )

    legacy_course_patches = {
        "Class 11–12 NEET": {
            "syllabus": ["Physics", "Chemistry", "Botany", "Zoology", "NCERT Mastery", "Weekly Mock Tests"],
            "faculty": ["Dr. A. Wani (Physics)", "Mr. R. Bhat (Chemistry)", "Ms. S. Kaur (Biology)"],
            "features": ["1:30 mentor ratio", "Doubt clearing daily", "AIIMS-style test series", "Personal performance dashboard"],
        },
        "IIT-JEE Main + Advanced": {
            "syllabus": ["Mathematics", "Physics", "Chemistry", "Numerical Practice", "Past JEE Papers"],
            "faculty": ["Mr. F. Lone (Maths)", "Dr. A. Wani (Physics)", "Mr. R. Bhat (Chemistry)"],
            "features": ["Small batches of 30", "Olympiad-grade problem sets", "All-India test ranking", "Doubt sessions 6 days a week"],
        },
    }
    for title, fields in legacy_course_patches.items():
        existing_course = await db.courses.find_one({"title": title})
        if existing_course:
            patch = {k: v for k, v in fields.items() if not existing_course.get(k)}
            if patch:
                await db.courses.update_one({"id": existing_course["id"]}, {"$set": patch})

    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)

async def _run_initial_seed():
    kashmir_centers = [
        {"id": new_id(), "name": "Unacademy Offline Centre Srinagar", "city": "Srinagar", "address": "Lal Chowk, Srinagar, J&K 190001", "phone": "+91-9876500001", "timing": "8:00 AM – 8:00 PM", "lat": 34.0837, "lng": 74.7973},
        {"id": new_id(), "name": "Unacademy Offline Centre Anantnag", "city": "Anantnag", "address": "KP Road, Anantnag, J&K 192101", "phone": "+91-9876500002", "timing": "8:00 AM – 8:00 PM", "lat": 33.7311, "lng": 75.1487},
        {"id": new_id(), "name": "Unacademy Offline Centre Sopore", "city": "Sopore", "address": "Main Chowk, Sopore, J&K 193201", "phone": "+91-9876500003", "timing": "8:00 AM – 8:00 PM", "lat": 34.2871, "lng": 74.4663},
        {"id": new_id(), "name": "Unacademy Offline Centre Soura", "city": "Soura", "address": "Soura, Srinagar, J&K 190011", "phone": "+91-9876500004", "timing": "8:00 AM – 8:00 PM", "lat": 34.1396, "lng": 74.8005},
        {"id": new_id(), "name": "Unacademy Offline Centre Zakura", "city": "Zakura", "address": "Zakura, Srinagar, J&K 190006", "phone": "+91-9876500005", "timing": "8:00 AM – 8:00 PM", "lat": 34.1373, "lng": 74.8584},
        {"id": new_id(), "name": "Unacademy Offline Centre Parraypora", "city": "Parraypora", "address": "Parraypora, Srinagar, J&K 190015", "phone": "+91-9876500006", "timing": "8:00 AM – 8:00 PM", "lat": 34.0500, "lng": 74.7833},
    ]
    await db.centers.insert_many(kashmir_centers)

    courses_data = [
        ("Class 11–12 NEET", "NEET", "24 months", 95000, "Comprehensive 2-year NEET preparation with Unacademy curriculum.", True, "https://images.unsplash.com/photo-1571260899304-425eee4c7efc?w=800",
         ["Physics", "Chemistry", "Botany", "Zoology", "NCERT Mastery", "Weekly Mock Tests"],
         ["Dr. A. Wani (Physics)", "Mr. R. Bhat (Chemistry)", "Ms. S. Kaur (Biology)"],
         ["1:30 mentor ratio", "Doubt clearing daily", "AIIMS-style test series", "Personal performance dashboard"]),
        ("IIT-JEE Main + Advanced", "IIT-JEE", "24 months", 105000, "Two-year integrated JEE Main + Advanced programme.", True, "https://images.pexels.com/photos/29534728/pexels-photo-29534728.jpeg?w=800",
         ["Mathematics", "Physics", "Chemistry", "Numerical Practice", "Past JEE Papers"],
         ["Mr. F. Lone (Maths)", "Dr. A. Wani (Physics)", "Mr. R. Bhat (Chemistry)"],
         ["Small batches of 30", "Olympiad-grade problem sets", "All-India test ranking", "Doubt sessions 6 days a week"]),
        ("Foundation 8th–10th", "Foundation", "12 months", 35000, "Strong academic foundation with Olympiad training.", False, "https://images.pexels.com/photos/6147219/pexels-photo-6147219.jpeg?w=800",
         ["Maths Foundation", "Science Foundation", "English", "Mental Ability", "Olympiad Prep"],
         ["Ms. M. Khan", "Mr. T. Rather"],
         ["Maths & Science foundation", "Olympiad level preparation", "Mental ability modules", "Small interactive batches"]),
        ("CBSE Class 11–12 Sciences", "CBSE", "24 months", 40000, "CBSE-aligned programme for PCM / PCB streams with Boards-grade rigor.", True, "https://images.unsplash.com/photo-1555967522-37949fc21dcb?w=800",
         ["NCERT Mastery", "Sample Paper Drills", "Practical Lab Notes", "Pre-Board Tests"],
         ["Mr. F. Lone", "Dr. A. Wani", "Ms. S. Kaur"],
         ["100% NCERT coverage", "Boards + competitive integration", "Pre-board mock series"]),
        ("JKBOSE 12th Boards", "JKBOSE", "10 months", 22000, "Targeted JKBOSE board syllabus mastery for Kashmir students.", False, "https://images.pexels.com/photos/29534728/pexels-photo-29534728.jpeg?w=800",
         ["JKBOSE Textbooks", "Weekly Topic Tests", "Previous Year Papers", "Viva Practice"],
         ["Local Faculty Panel"],
         ["JKBOSE-pattern test series", "One-on-one revision plans", "Affordable monthly fee plans"]),
    ]
    for t, cat, dur, fee, desc, feat, img, syl, fac, features in courses_data:
        await db.courses.insert_one({
            "id": new_id(), "title": t, "category": cat, "duration": dur, "fee": fee,
            "description": desc, "syllabus": syl, "faculty": fac, "features": features,
            "scholarship_available": True, "featured": feat, "image_url": img, "created_at": now_iso(),
        })

    notices = [
        ("New 2026 NEET Batch Launch", "Admissions open for the new NEET 2026 batch starting March 1.", "Admissions", True),
        ("Scholarship Test 2026", "Unacademy Offline Scholarship Test on Feb 28 — up to 100% fee waiver.", "Scholarship", True),
        ("Foundation Olympiad Workshop", "Free 3-day Olympiad workshop for Class 8–10 students.", "Workshop", False),
    ]
    for t, c, cat, p in notices:
        await db.notices.insert_one({"id": new_id(), "title": t, "content": c, "category": cat, "pinned": p, "created_at": now_iso()})

    results_data = [
        ("Aamir Hussain", "NEET 2025", "AIR 412", 2025, "NEET", "Cracked NEET in first attempt with guidance."),
        ("Zoya Bhat", "JEE Advanced 2025", "AIR 1108", 2025, "IIT-JEE", "From Anantnag to IIT Delhi — mentors made the difference."),
        ("Hamid Wani", "NEET 2024", "AIR 587", 2024, "NEET", "Dedicated faculty + structured tests = AIIMS dream realised."),
        ("Sahla Mir", "CUET 2024", "99.4 percentile", 2024, "CUET", "Got admission into Delhi University Hindu College."),
        ("Bilal Ahmad", "JEE Main 2025", "99.1 percentile", 2025, "IIT-JEE", "Best teaching ecosystem in Kashmir, hands down."),
        ("Iqra Jan", "NEET 2025", "AIR 1903", 2025, "NEET", "Made the impossible feel routine."),
    ]
    for n, e, r, y, c, q in results_data:
        await db.results.insert_one({
            "id": new_id(), "student_name": n, "exam": e, "rank": r, "year": y, "course": c,
            "photo_url": None, "quote": q, "created_at": now_iso()
        })

    ts = [
        ("Insha Rather", "Parent", "Transformed my daughter's preparation. The faculty truly cares."),
        ("Rayaan Khan", "NEET Aspirant", "Best decision was joining in Srinagar. Mock tests were spot on."),
        ("Mehak Lone", "JEE Aspirant", "Doubt clearing happens in real time — feels like national coaching."),
    ]
    for n, role, q in ts:
        await db.testimonials.insert_one({"id": new_id(), "name": n, "role": role, "quote": q, "created_at": now_iso()})

    jobs = [
        ("Physics Faculty (NEET/JEE)", "Academics", "Srinagar", "Full-time", "Senior physics educator for NEET/JEE batches.", ["M.Sc/Ph.D Physics", "3+ years coaching experience"], True),
        ("Counselor", "Admissions", "Anantnag", "Full-time", "Student counseling and parent interactions.", ["Graduate", "Excellent communication"], True),
        ("Floor Manager", "Operations", "Sopore", "Full-time", "Manage center operations and student discipline.", ["Graduate", "Leadership skills"], True),
        ("BDM (Business Dev. Manager)", "Business", "Srinagar", "Full-time", "Drive admissions and outreach across Kashmir.", ["MBA preferred", "5+ years in EdTech"], True),
        ("DTP Operator", "Production", "Srinagar", "Full-time", "Design study material and notices.", ["CorelDraw / InDesign expertise"], True),
    ]
    for t, d, l, ty, desc, req, a in jobs:
        await db.jobs.insert_one({"id": new_id(), "title": t, "department": d, "location": l, "type": ty, "description": desc, "requirements": req, "active": a, "created_at": now_iso()})

    await db.scholarships.insert_one({
        "id": "1ed94009-d949-4504-b291-68e3571a5a44", "title": "Unacademy Offline Centre Scholarship Test 2026",
        "description": "Win up to 100% scholarship on tuition fees. Open for Class 8–12 students across Kashmir.",
        "exam_date": "2026-02-28", "deadline": "2026-02-25",
        "eligibility": "Students of Class 8 to 12 from any school in J&K.",
        "active": True,
        "examiner_token": uuid.uuid4().hex,
        "available_venues": ["90 FT", "Anantnag", "Sopore", "Zakura", "Parraypora"],
        "created_at": now_iso(),
    })

# ---------- App wiring ----------
from erp_routes import build_erp_router, erp_seed
erp_router = build_erp_router(db, get_current_user, hash_password, verify_password, require_admin)
api.include_router(erp_router)

from whatsapp_inbox import build_whatsapp_router
api.include_router(build_whatsapp_router(db, require_super_admin))
# WATH Carnival + page config
from wath_carnival import build_wath_router, try_reserve_slot, release_slot  # noqa: E402
api.include_router(build_wath_router(db, require_admin))
app.include_router(api)

_default_allowed = [
    "http://localhost:3000",
    "https://nexed-neet.preview.emergentagent.com",
    "https://northendedu.com",
    "https://www.northendedu.com",
    "https://nexed-neet.emergent.host",
]
_extra = os.environ.get("ADDITIONAL_ORIGINS", "")
if _extra:
    _default_allowed.extend([o.strip() for o in _extra.split(",") if o.strip()])
_frontend_url = os.environ.get("FRONTEND_URL", "")
if _frontend_url and _frontend_url not in _default_allowed:
    _default_allowed.append(_frontend_url)

app.add_middleware(
    CORSMiddleware,
    allow_origins=list(set(_default_allowed)),
    allow_origin_regex=r"https://([a-z0-9-]+\.)?(preview\.emergentagent\.com|emergent\.host|northendedu\.com)$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)

@app.get("/health")
async def health():
    return {"status": "ok"}

@app.get("/")
async def root():
    return {"status": "ok", "service": "Unacademy Offline Centre API"}

async def _run_boot_tasks():
    try:
        await seed()
    except Exception as e:
        logging.error(f"seed() failed: {e}")
    try:
        await erp_seed(db, hash_password)
    except Exception as e:
        logging.error(f"erp_seed() failed: {e}")
    try:
        await init_storage()
    except Exception as e:
        logging.error(f"init_storage() failed: {e}")
    try:
        await _backfill_slugs()
    except Exception as e:
        logging.error(f"_backfill_slugs() failed: {e}")
    logging.info("Unacademy Offline Centre backend ready.")


async def _backfill_slugs():
    for coll in ("courses", "scholarships"):
        cursor = db[coll].find({"$or": [{"slug": {"$exists": False}}, {"slug": None}, {"slug": ""}]}, {"id": 1, "title": 1})
        async for doc in cursor:
            slug = await unique_slug(coll, doc.get("title") or "item", exclude_id=doc.get("id"))
            await db[coll].update_one({"id": doc["id"]}, {"$set": {"slug": slug}})

@app.on_event("startup")
async def on_start():
    asyncio.create_task(_run_boot_tasks())

@app.on_event("shutdown")
async def on_stop():
    await storage_aclose()
    client.close()